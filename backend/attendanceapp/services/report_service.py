"""
Daily Excel report generator.
"""

from pathlib import Path

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from .attendance_processor import format_duration


def _time_value(value):
    if not value:
        return ""
    return timezone.localtime(value).strftime("%H:%M:%S")


def _is_week_off_present(row):
    return row.attendance_date.weekday() >= 5 and bool(
        row.total_punches or row.punch_in or row.punch_out
    )


def generate_excel_report(attendance_rows, report_date):
    rows = list(attendance_rows)
    reports_dir = Path(settings.MEDIA_ROOT) / "attendance" / "reports" / "daily"
    reports_dir.mkdir(parents=True, exist_ok=True)

    filename = f"Daily_Attendance_Report_{report_date:%Y-%m-%d}.xlsx"
    report_path = reports_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    headers = [
        "Employee Code",
        "Employee Name",
        "Date",
        "Punch In",
        "Punch Out",
        "Gross Hours",
        "Break Time",
        "Net Hours",
        "Late Status",
        "Early Exit",
        "Missing Punch",
        "Week Off Present",
        "Status",
    ]
    ws.append(headers)
    ws["G1"].comment = Comment(
        "Break Time is inferred from intermediate punches because eSSL transaction logs do not include explicit break type.",
        "eSSL Attendance",
    )

    header_fill = PatternFill("solid", fgColor="1C2744")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        is_week_off_present = _is_week_off_present(row)
        status = "Week Off Present" if is_week_off_present else row.get_status_display()
        ws.append(
            [
                row.employee_code,
                row.employee_name,
                row.attendance_date.strftime("%Y-%m-%d"),
                _time_value(row.punch_in),
                _time_value(row.punch_out),
                format_duration(row.working_hours_seconds),
                format_duration(row.break_time_seconds),
                format_duration(row.net_working_hours_seconds),
                "Late" if row.is_late else "On Time",
                "Yes" if row.is_early_exit else "No",
                "Yes" if row.missing_punch else "No",
                "Yes" if is_week_off_present else "No",
                status,
            ]
        )

    for row_cells in ws.iter_rows(min_row=2):
        late_cell = row_cells[8]
        missing_cell = row_cells[10]
        week_off_present_cell = row_cells[11]
        status_cell = row_cells[12]
        if late_cell.value == "Late":
            late_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        if missing_cell.value == "Yes":
            missing_cell.fill = PatternFill("solid", fgColor="F4CCCC")
        if week_off_present_cell.value == "Yes":
            week_off_present_cell.fill = PatternFill("solid", fgColor="D9EAD3")
        if status_cell.value == "Present":
            status_cell.fill = PatternFill("solid", fgColor="D9EAD3")

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 3, 36)

    ws.freeze_panes = "A2"
    wb.save(report_path)
    return f"attendance/reports/daily/{filename}", report_path