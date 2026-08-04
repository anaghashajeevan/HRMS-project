"""
Monthly attendance report service.
Adapted for HRMS integration: pulls active employees from HRMSapp.Employee.
"""

import calendar
import logging
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path

from django.conf import settings
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 🎯 HRMS INTEGRATION: Import HRMS Employee instead of attendance Employee
from HRMSapp.models import Employee

from attendanceapp.models import (
    AutomationSettings,
    DailyAttendance,
    MonthlyReportLog,
    normalize_employee_code,
)

from .attendance_processor import format_duration, process_attendance_period
from .email_service import parse_recipients, resolve_monthly_receiver, send_monthly_attendance_report
from .essl_service import call_essl_api_for_range, extract_str_data_list, parse_punch_logs

logger = logging.getLogger(__name__)


class MonthlyAttendanceError(Exception):
    pass


@dataclass(frozen=True)
class MonthlyPeriod:
    year: int
    month: int
    start_date: date
    end_date: date
    from_datetime: datetime
    to_datetime: datetime
    included_dates: tuple[date, ...]

    @property
    def monthly_days(self):
        return len(self.included_dates)


def month_period(year, month, today=None):
    start_date = date(year, month, 1)
    end_date = date(year, month, calendar.monthrange(year, month)[1])
    return start_date, end_date


def _parse_excluded_dates(value):
    excluded = set()
    for item in re.split(r"[\s,;]+", value or ""):
        if not item:
            continue
        for date_format in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                excluded.add(datetime.strptime(item, date_format).date())
                break
            except ValueError:
                continue
    return excluded


def resolve_report_period(year, month, settings_obj=None, today=None):
    settings_obj = settings_obj or AutomationSettings.get_solo()
    start_date, end_date = month_period(year, month, today=today)
    if settings_obj.monthly_report_start_date and settings_obj.monthly_report_end_date:
        start_date = settings_obj.monthly_report_start_date
        end_date = settings_obj.monthly_report_end_date
    elif settings_obj.monthly_report_start_date:
        start_date = settings_obj.monthly_report_start_date
    elif settings_obj.monthly_report_end_date:
        end_date = settings_obj.monthly_report_end_date

    if start_date > end_date:
        raise MonthlyAttendanceError("Report Start Date must be before Report End Date.")

    excluded_dates = _parse_excluded_dates(settings_obj.excluded_dates)
    included_dates = []
    current = start_date
    while current <= end_date:
        if current not in excluded_dates:
            included_dates.append(current)
        current += timedelta(days=1)

    return MonthlyPeriod(
        year=year,
        month=month,
        start_date=start_date,
        end_date=end_date,
        from_datetime=datetime.combine(start_date, time.min),
        to_datetime=datetime.combine(end_date, time(23, 59, 59)),
        included_dates=tuple(included_dates),
    )


def month_datetime_period(year, month, today=None):
    return resolve_report_period(year, month, today=today)


def _previous_month(value):
    if value.month == 1:
        return value.year - 1, 12
    return value.year, value.month - 1


def target_report_month(settings_obj, today=None):
    today = today or timezone.localdate()
    if settings_obj.monthly_report_start_date:
        return settings_obj.monthly_report_start_date.year, settings_obj.monthly_report_start_date.month
    if settings_obj.monthly_report_mode == AutomationSettings.MONTHLY_MODE_CURRENT:
        return today.year, today.month
    return _previous_month(today)


def _seconds_from_hours(hours):
    return int(float(hours or 0) * 3600)


def _time_value(value):
    if not value:
        return "-"
    return timezone.localtime(value).strftime("%H:%M:%S")


def _display_period(period):
    return f"{period.start_date:%d-%b-%Y} to {period.end_date:%d-%b-%Y}"


def _date_range_label(start_date, end_date):
    return f"{start_date:%d-%b-%Y} to {end_date:%d-%b-%Y}"


def _is_week_off(value):
    return value.weekday() >= 5


def _period_day_counts(period):
    week_off_days = sum(1 for item in period.included_dates if _is_week_off(item))
    return {
        "week_off_days": week_off_days,
        "working_days": period.monthly_days - week_off_days,
    }


def _local_time(value):
    return timezone.localtime(value).time() if value else None


def _setting_time(value):
    if isinstance(value, time):
        return value
    return datetime.strptime(str(value), "%H:%M").time()


def _lunch_overlap_seconds(row, settings_obj):
    if not row.punch_in or not row.punch_out:
        return 0
    local_in = timezone.localtime(row.punch_in)
    local_out = timezone.localtime(row.punch_out)
    lunch_start = timezone.make_aware(
        datetime.combine(local_in.date(), _setting_time(settings_obj.lunch_start_time)),
        timezone.get_default_timezone(),
    )
    lunch_end = timezone.make_aware(
        datetime.combine(local_in.date(), _setting_time(settings_obj.lunch_end_time)),
        timezone.get_default_timezone(),
    )
    overlap_start = max(local_in, lunch_start)
    overlap_end = min(local_out, lunch_end)
    if overlap_end <= overlap_start:
        return 0
    return int((overlap_end - overlap_start).total_seconds())


def _metrics_for_row(row, settings_obj):
    if not row or row.missing_punch or not row.punch_in or not row.punch_out:
        return {
            "gross_seconds": 0,
            "break_seconds": 0,
            "net_seconds": 0,
            "status": "Missing Punch" if row else "Absent",
            "low_working_hours": False,
        }

    gross_seconds = row.working_hours_seconds
    break_seconds = max(row.break_time_seconds, _lunch_overlap_seconds(row, settings_obj))
    net_seconds = max(gross_seconds - break_seconds, 0)
    full_day_seconds = _seconds_from_hours(settings_obj.full_day_min_hours)
    punch_out_time = _local_time(row.punch_out)

    is_present = bool(
        punch_out_time
        and punch_out_time >= _setting_time(settings_obj.full_day_out_time)
        or net_seconds >= full_day_seconds
    )
    low_working_hours = not is_present

    return {
        "gross_seconds": gross_seconds,
        "break_seconds": break_seconds,
        "net_seconds": net_seconds,
        "status": "Present",
        "low_working_hours": low_working_hours,
    }


def fetch_monthly_logs(year, month, settings_obj=None, today=None):
    settings_obj = settings_obj or AutomationSettings.get_solo()
    period = resolve_report_period(year, month, settings_obj=settings_obj, today=today)
    logger.info(
        "Fetching monthly eSSL logs: start_date=%s end_date=%s",
        period.start_date,
        period.end_date,
    )
    xml_text = call_essl_api_for_range(
        settings_obj,
        period.from_datetime,
        period.to_datetime,
    )
    _, str_data = extract_str_data_list(xml_text)
    if not str_data:
        raise MonthlyAttendanceError("No punch logs received from eSSL API for the selected report period.")
    logs = parse_punch_logs(str_data)
    if not logs:
        raise MonthlyAttendanceError("No valid punch logs found in the monthly eSSL response.")
    return logs, period


def process_monthly_logs(year, month, settings_obj=None, today=None, logs=None):
    settings_obj = settings_obj or AutomationSettings.get_solo()
    if logs is None:
        logs, period = fetch_monthly_logs(year, month, settings_obj=settings_obj, today=today)
    else:
        period = resolve_report_period(year, month, settings_obj=settings_obj, today=today)

    attendance_rows = process_attendance_period(
        logs,
        settings_obj,
        period.start_date,
        period.end_date,
    )
    return attendance_rows, period


def _hrms_employee_display(hrms_employee):
    """Get display info from HRMS Employee."""
    department = ""
    if hrms_employee.department:
        department = hrms_employee.department.name
    elif hrms_employee.structure_location:
        department = hrms_employee.structure_location.name
    return {
        "employee_code": hrms_employee.employee_id,
        "employee_name": hrms_employee.full_name,
        "department": department,
    }


def _blank_summary_for_hrms_employee(hrms_employee, period):
    day_counts = _period_day_counts(period)
    info = _hrms_employee_display(hrms_employee)
    return {
        "employee_code": info["employee_code"],
        "employee_name": info["employee_name"],
        "department": info["department"],
        "monthly_days": period.monthly_days,
        "week_off_days": day_counts["week_off_days"],
        "week_off_present_days": 0,
        "working_days": day_counts["working_days"],
        "attendance_days": 0,
        "absent_days": 0,
        "missing_punch_days": 0,
        "total_gross_seconds": 0,
        "total_break_seconds": 0,
        "total_working_seconds": 0,
    }


def collect_monthly_attendance(year, month, settings_obj=None, today=None, attendance_rows=None):
    """
    Aggregate monthly attendance stats per employee + department.
    Uses HRMS Employees as the master list.
    """
    settings_obj = settings_obj or AutomationSettings.get_solo()
    period = resolve_report_period(year, month, settings_obj=settings_obj, today=today)
    day_counts = _period_day_counts(period)
    rows = list(attendance_rows) if attendance_rows is not None else list(
        DailyAttendance.objects.filter(
            attendance_date__gte=period.start_date,
            attendance_date__lte=period.end_date,
        )
    )
    rows_by_key = {(row.employee_code, row.attendance_date): row for row in rows}

    # 🎯 HRMS INTEGRATION: Use HRMS Employees as master list
    active_employees = list(
        Employee.objects.filter(
            is_deleted=False,
            status__in=['ACTIVE', 'PROBATION'],
        ).select_related('structure_location', 'department', 'location').order_by('employee_id')
    )
    # Build a set of normalized HRMS codes for quick lookup
    active_codes_normalized = {
        normalize_employee_code(emp.employee_id): emp for emp in active_employees
    }

    logger.info(
        "Collecting monthly attendance: employee_master_count=%s start_date=%s end_date=%s monthly_days=%s",
        len(active_employees),
        period.start_date,
        period.end_date,
        period.monthly_days,
    )

    employee_summaries = {
        emp.employee_id: _blank_summary_for_hrms_employee(emp, period)
        for emp in active_employees
    }
    department_summaries = defaultdict(
        lambda: {
            "department": "",
            "employees": set(),
            "monthly_days": period.monthly_days,
            "week_off_days": day_counts["week_off_days"],
            "week_off_present_days": 0,
            "working_days": day_counts["working_days"],
            "attendance_days": 0,
            "absent_days": 0,
            "missing_punch_days": 0,
            "total_gross_seconds": 0,
            "total_break_seconds": 0,
            "total_working_seconds": 0,
        }
    )
    daily_details = []
    exceptions = []

    for employee in active_employees:
        info = _hrms_employee_display(employee)
        department = info["department"] or "Unassigned"
        department_summary = department_summaries[department]
        department_summary["department"] = department
        department_summary["employees"].add(employee.employee_id)
        employee_summary = employee_summaries[employee.employee_id]

        # Match rows by normalized code
        emp_code_normalized = normalize_employee_code(employee.employee_id)

        for attendance_date in period.included_dates:
            # Find row matching this employee (by normalized code)
            row = None
            for row_key_code, row_key_date in rows_by_key.keys():
                if row_key_date == attendance_date and normalize_employee_code(row_key_code) == emp_code_normalized:
                    row = rows_by_key[(row_key_code, row_key_date)]
                    break

            metrics = _metrics_for_row(row, settings_obj)
            is_week_off = _is_week_off(attendance_date)
            has_punch = bool(row)
            if is_week_off and has_punch:
                status = "Week Off Present"
            elif is_week_off:
                status = "Week Off"
            else:
                status = metrics["status"]

            if status in {"Present", "Week Off Present"}:
                employee_summary["attendance_days"] += 1
                department_summary["attendance_days"] += 1
                if status == "Week Off Present":
                    employee_summary["week_off_present_days"] += 1
                    department_summary["week_off_present_days"] += 1
            elif status == "Missing Punch":
                employee_summary["missing_punch_days"] += 1
                department_summary["missing_punch_days"] += 1
            elif status == "Absent":
                employee_summary["absent_days"] += 1
                department_summary["absent_days"] += 1

            if status in {"Present", "Week Off Present"}:
                employee_summary["total_gross_seconds"] += metrics["gross_seconds"]
                employee_summary["total_break_seconds"] += metrics["break_seconds"]
                employee_summary["total_working_seconds"] += metrics["net_seconds"]
                department_summary["total_gross_seconds"] += metrics["gross_seconds"]
                department_summary["total_break_seconds"] += metrics["break_seconds"]
                department_summary["total_working_seconds"] += metrics["net_seconds"]

            remarks = _remarks_for_detail(status, metrics)
            day_name = attendance_date.strftime("%A")

            detail = {
                "date": attendance_date,
                "day": day_name,
                "employee_code": info["employee_code"],
                "employee_name": info["employee_name"],
                "department": info["department"],
                "punch_in": _time_value(row.punch_in) if row else "-",
                "punch_out": _time_value(row.punch_out) if row else "-",
                "gross_hours": format_duration(metrics["gross_seconds"]),
                "break_time": format_duration(metrics["break_seconds"]),
                "working_hours": format_duration(metrics["net_seconds"]),
                "status": status,
                "remarks": remarks,
            }
            daily_details.append(detail)

            issue_types = []
            if status in {"Absent", "Missing Punch", "Week Off Present"}:
                issue_types.append(status)
            if metrics["low_working_hours"] and status in {"Present", "Week Off Present"}:
                issue_types.append("Low Working Hours")
            for issue_type in issue_types:
                exceptions.append(
                    {
                        **detail,
                        "issue_type": issue_type,
                        "remarks": _remarks_for_issue(issue_type),
                    }
                )

    # Handle unknown employees (in DailyAttendance but not in HRMS)
    unknown_rows = []
    for row in rows:
        row_code_normalized = normalize_employee_code(row.employee_code)
        if row_code_normalized not in active_codes_normalized and row.attendance_date in period.included_dates:
            unknown_rows.append(row)

    for row in sorted(unknown_rows, key=lambda item: (item.attendance_date, item.employee_code)):
        metrics = _metrics_for_row(row, settings_obj)
        status = "Week Off Present" if _is_week_off(row.attendance_date) else metrics["status"]
        detail = {
            "date": row.attendance_date,
            "day": row.attendance_date.strftime("%A"),
            "employee_code": row.employee_code,
            "employee_name": row.employee_name,
            "department": "",
            "punch_in": _time_value(row.punch_in),
            "punch_out": _time_value(row.punch_out),
            "gross_hours": format_duration(metrics["gross_seconds"]),
            "break_time": format_duration(metrics["break_seconds"]),
            "working_hours": format_duration(metrics["net_seconds"]),
            "status": status,
            "remarks": "Employee code was not found in active HRMS employees.",
        }
        daily_details.append(detail)
        exceptions.append(
            {
                **detail,
                "issue_type": "Unknown Employee",
                "remarks": "Employee code was not found in active HRMS employees.",
            }
        )

    department_rows = []
    for department in sorted(department_summaries):
        item = department_summaries[department]
        department_rows.append(
            {
                "department": item["department"],
                "employees_count": len(item["employees"]),
                "monthly_days": item["monthly_days"],
                "week_off_days": item["week_off_days"],
                "week_off_present_days": item["week_off_present_days"],
                "working_days": item["working_days"],
                "attendance_days": item["attendance_days"],
                "absent_days": item["absent_days"],
                "missing_punch_days": item["missing_punch_days"],
                "total_gross_seconds": item["total_gross_seconds"],
                "total_break_seconds": item["total_break_seconds"],
                "total_working_seconds": item["total_working_seconds"],
            }
        )

    totals = {
        "employees_count": len(active_employees),
        "monthly_days": period.monthly_days,
        "week_off_days": day_counts["week_off_days"],
        "week_off_present_days": sum(item["week_off_present_days"] for item in employee_summaries.values()),
        "working_days": day_counts["working_days"],
        "attendance_days": sum(item["attendance_days"] for item in employee_summaries.values()),
        "absent_days": sum(item["absent_days"] for item in employee_summaries.values()),
        "missing_punch_days": sum(item["missing_punch_days"] for item in employee_summaries.values()),
        "total_gross_seconds": sum(item["total_gross_seconds"] for item in employee_summaries.values()),
        "total_break_seconds": sum(item["total_break_seconds"] for item in employee_summaries.values()),
        "total_working_seconds": sum(item["total_working_seconds"] for item in employee_summaries.values()),
    }
    totals["total_hours"] = format_duration(totals["total_gross_seconds"])
    totals["total_break_time"] = format_duration(totals["total_break_seconds"])
    totals["total_working_hours"] = format_duration(totals["total_working_seconds"])

    return {
        "year": year,
        "month": month,
        "month_label": date(year, month, 1).strftime("%B %Y"),
        "start_date": period.start_date,
        "end_date": period.end_date,
        "period_label": _display_period(period),
        "monthly_days": period.monthly_days,
        "employee_summaries": list(employee_summaries.values()),
        "daily_details": sorted(daily_details, key=lambda item: (item["date"], item["employee_code"])),
        "department_summaries": department_rows,
        "exceptions": sorted(exceptions, key=lambda item: (item["date"], item["employee_code"], item["issue_type"])),
        "totals": totals,
    }


def _remarks_for_issue(issue_type):
    return {
        "Absent": "No punch on working day.",
        "Missing Punch": "Only one punch found.",
        "Week Off Present": "Worked on week off.",
        "Unknown Employee": "Employee code was not found in active HRMS employees.",
        "Low Working Hours": "Low working hours.",
    }.get(issue_type, "")


def _remarks_for_detail(status, metrics):
    remarks = []
    if status == "Week Off Present":
        remarks.append("Worked on week off")
    elif status == "Absent":
        remarks.append("No punch on working day")
    elif status == "Missing Punch":
        remarks.append("Only one punch found")

    if metrics["low_working_hours"] and status in {"Present", "Week Off Present"}:
        remarks.append("Low working hours")
    return "; ".join(remarks)


def _append_report_header(sheet, report):
    last_column = get_column_letter(max(sheet.max_column, 1))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = "Company: NL Technology"
    sheet["A1"].font = Font(bold=True, size=15, color="1C2744")
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = "Monthly Attendance Report"
    sheet["A2"].font = Font(bold=True, size=14, color="1C2744")
    sheet["A3"] = f"Report Period: {_date_range_label(report['start_date'], report['end_date'])}"
    sheet["A4"] = f"Generated On: {timezone.localtime(timezone.now()):%d-%b-%Y %H:%M:%S}"
    sheet["A3"].font = Font(bold=True, color="475569")
    sheet["A4"].font = Font(color="475569")


def _style_table(sheet, header_row=6, highlight_from_col=None):
    header_fill = PatternFill("solid", fgColor="1C2744")
    header_font = Font(bold=True, color="FFFFFF")
    light_fill = PatternFill("solid", fgColor="EEF4FF")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for cell in sheet[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if highlight_from_col and cell.column >= highlight_from_col:
                cell.fill = light_fill


def _autosize(sheet):
    for column_index, column in enumerate(sheet.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(column_index)
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 38)


def _prepare_sheet(sheet, report, highlight_from_col=None):
    _append_report_header(sheet, report)
    _style_table(sheet, header_row=6, highlight_from_col=highlight_from_col)
    _autosize(sheet)
    sheet.freeze_panes = "A7"


def generate_monthly_excel_report(year, month, settings_obj=None, today=None, logs=None):
    settings_obj = settings_obj or AutomationSettings.get_solo()
    attendance_rows, _ = process_monthly_logs(
        year,
        month,
        settings_obj=settings_obj,
        today=today,
        logs=logs,
    )
    report = collect_monthly_attendance(
        year,
        month,
        settings_obj=settings_obj,
        today=today,
        attendance_rows=attendance_rows,
    )
    reports_dir = Path(settings.MEDIA_ROOT) / "attendance" / "reports" / "monthly"
    reports_dir.mkdir(parents=True, exist_ok=True)

    filename = f"Monthly_Attendance_Report_{year}-{month:02d}.xlsx"
    report_path = reports_dir / filename

    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Monthly Summary"
    for _ in range(5):
        summary_sheet.append([])
    summary_sheet.append(
        [
            "Employee Code",
            "Employee Name",
            "Department",
            "Monthly Days",
            "Week Off Days",
            "Week Off Present Days",
            "Working Days",
            "Present Days",
            "Absent Days",
            "Missing Punch Days",
            "Total Hours",
            "Total Break Time",
            "Total Working Hours",
        ]
    )
    for row in report["employee_summaries"]:
        summary_sheet.append(
            [
                row["employee_code"],
                row["employee_name"],
                row["department"],
                row["monthly_days"],
                row["week_off_days"],
                row["week_off_present_days"],
                row["working_days"],
                row["attendance_days"],
                row["absent_days"],
                row["missing_punch_days"],
                format_duration(row["total_gross_seconds"]),
                format_duration(row["total_break_seconds"]),
                format_duration(row["total_working_seconds"]),
            ]
        )

    details_sheet = wb.create_sheet("Daily Punch Details")
    for _ in range(5):
        details_sheet.append([])
    details_sheet.append(
        [
            "Date",
            "Day",
            "Week Off Present",
            "Employee Code",
            "Employee Name",
            "Department",
            "Punch In",
            "Punch Out",
            "Total Hours",
            "Break Time",
            "Working Hours",
            "Status",
            "Remarks",
        ]
    )
    details_sheet["J6"].comment = Comment(
        "Break Time uses inferred punches and at least the configured lunch window when applicable.",
        "eSSL Attendance",
    )
    for row in report["daily_details"]:
        details_sheet.append(
            [
                row["date"].strftime("%Y-%m-%d"),
                row["day"],
                "Yes" if row["status"] == "Week Off Present" else "No",
                row["employee_code"],
                row["employee_name"],
                row["department"],
                row["punch_in"],
                row["punch_out"],
                row["gross_hours"],
                row["break_time"],
                row["working_hours"],
                row["status"],
                row["remarks"],
            ]
        )

    department_sheet = wb.create_sheet("Department Summary")
    for _ in range(5):
        department_sheet.append([])
    department_sheet.append(
        [
            "Department",
            "Employees Count",
            "Monthly Days",
            "Week Off Days",
            "Week Off Present Days",
            "Working Days",
            "Attendance Days",
            "Absent Days",
            "Missing Punch Days",
            "Total Hours",
            "Total Break Time",
            "Total Working Hours",
        ]
    )
    for row in report["department_summaries"]:
        department_sheet.append(
            [
                row["department"],
                row["employees_count"],
                row["monthly_days"],
                row["week_off_days"],
                row["week_off_present_days"],
                row["working_days"],
                row["attendance_days"],
                row["absent_days"],
                row["missing_punch_days"],
                format_duration(row["total_gross_seconds"]),
                format_duration(row["total_break_seconds"]),
                format_duration(row["total_working_seconds"]),
            ]
        )

    exceptions_sheet = wb.create_sheet("Exceptions")
    for _ in range(5):
        exceptions_sheet.append([])
    exceptions_sheet.append(
        [
            "Date",
            "Day",
            "Employee Code",
            "Employee Name",
            "Department",
            "Issue Type",
            "Punch In",
            "Punch Out",
            "Working Hours",
            "Remarks",
        ]
    )
    for row in report["exceptions"]:
        exceptions_sheet.append(
            [
                row["date"].strftime("%Y-%m-%d"),
                row["day"],
                row["employee_code"],
                row["employee_name"],
                row["department"],
                row["issue_type"],
                row["punch_in"],
                row["punch_out"],
                row["working_hours"],
                row["remarks"],
            ]
        )

    _prepare_sheet(summary_sheet, report, highlight_from_col=4)
    _prepare_sheet(details_sheet, report)
    _prepare_sheet(department_sheet, report, highlight_from_col=3)
    _prepare_sheet(exceptions_sheet, report)

    wb.save(report_path)
    logger.info("Monthly report generated: %s", report_path)
    return report, f"attendance/reports/monthly/{filename}", report_path


def generate_monthly_report(year, month, created_by=None):
    settings_obj = AutomationSettings.get_solo()
    report, report_file_name, _ = generate_monthly_excel_report(
        year,
        month,
        settings_obj=settings_obj,
    )
    log, _ = MonthlyReportLog.objects.get_or_create(
        year=year,
        month=month,
        defaults={
            "status": MonthlyReportLog.STATUS_GENERATED,
            "created_by": created_by,
        },
    )
    if created_by and not log.created_by:
        log.created_by = created_by
    log.mark_generated(report_file_name)
    return log, report


def send_monthly_report(year, month, created_by=None, prevent_duplicate=True, manual_send=False):
    settings_obj = AutomationSettings.get_solo()
    log = MonthlyReportLog.objects.filter(year=year, month=month).first()
    if not settings_obj.enable_monthly_report and not manual_send:
        if not log:
            log = MonthlyReportLog.objects.create(
                year=year,
                month=month,
                status=MonthlyReportLog.STATUS_FAILED,
                created_by=created_by,
                error_message="Monthly Report Email is disabled.",
            )
        return {
            "ok": False,
            "skipped": True,
            "message": "Monthly Report Email is disabled.",
            "log": log,
            "report": None,
        }
    if prevent_duplicate and log and log.status == MonthlyReportLog.STATUS_SUCCESS:
        return {
            "ok": True,
            "skipped": True,
            "message": "Monthly attendance report has already been sent.",
            "log": log,
            "report": collect_monthly_attendance(year, month, settings_obj=settings_obj),
        }

    log, _ = MonthlyReportLog.objects.get_or_create(
        year=year,
        month=month,
        defaults={
            "status": MonthlyReportLog.STATUS_GENERATED,
            "created_by": created_by,
        },
    )
    if created_by and not log.created_by:
        log.created_by = created_by
        log.save(update_fields=["created_by"])

    try:
        report, report_file_name, report_path = generate_monthly_excel_report(
            year,
            month,
            settings_obj=settings_obj,
        )
        with transaction.atomic():
            log.mark_generated(report_file_name)
    except Exception as exc:
        log.mark_failed(str(exc) or "Monthly attendance report generation failed.")
        return {
            "ok": False,
            "skipped": False,
            "message": str(exc) or "Monthly attendance report generation failed.",
            "log": log,
            "report": None,
        }

    try:
        sent_to, cc = send_monthly_attendance_report(
            settings_obj,
            report["period_label"],
            report["totals"],
            report_path,
        )
        log.mark_success(report_file_name, sent_to, ", ".join(cc))
        return {
            "ok": True,
            "skipped": False,
            "message": "Monthly attendance report sent successfully.",
            "log": log,
            "report": report,
        }
    except Exception as exc:
        log.report_file.name = report_file_name
        log.mark_failed(str(exc) or "Monthly attendance email failed.")
        return {
            "ok": False,
            "skipped": False,
            "message": str(exc) or "Monthly attendance email failed.",
            "log": log,
            "report": report,
        }


def run_monthly_attendance_automation(created_by=None, force_send=True):
    settings_obj = AutomationSettings.get_solo()
    year, month = target_report_month(settings_obj)
    steps = [
        "Connecting to eSSL API",
        "Fetching monthly punch logs",
        "Processing monthly attendance",
        "Generating monthly Excel",
        "Sending monthly email",
    ]
    result = send_monthly_report(
        year,
        month,
        created_by=created_by,
        prevent_duplicate=not force_send,
        manual_send=force_send,
    )
    if result["ok"]:
        steps.append("Completed")
    return {**result, "steps": steps}


def monthly_recipients(settings_obj):
    receiver, _ = resolve_monthly_receiver(settings_obj)
    return (
        receiver,
        parse_recipients(settings_obj.monthly_cc_emails),
    )