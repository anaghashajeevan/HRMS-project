from __future__ import annotations

from calendar import month_name
from decimal import Decimal
from io import BytesIO

from django.core.exceptions import ObjectDoesNotExist
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from xml.sax.saxutils import escape

from .models import (
    ExpenseAttachment,
    ExpenseItem,
    GeneratedReport,
    MonthlyReimbursementBatch,
    ReimbursementClaim,
    SmartReimbursementUpload,
    get_system_settings,
)


def _generated_timestamp() -> str:
    return timezone.localtime(timezone.now()).strftime("%d %b %Y %I:%M %p")


def _money(value: Decimal | int | float | str) -> str:
    return f"{Decimal(str(value or 0)).quantize(Decimal('0.01')):.2f}"


def _expense_bill_status(expense_item: ExpenseItem) -> str:
    attachments = getattr(expense_item, "prefetched_attachments", None)
    if attachments is None:
        attachments = list(expense_item.attachments.select_related("extraction").all())
    return "Bill Attached" if attachments else "Missing Bill"


def _expense_ocr_status(expense_item: ExpenseItem) -> str:
    attachments = getattr(expense_item, "prefetched_attachments", None)
    if attachments is None:
        attachments = list(expense_item.attachments.select_related("extraction").all())
    if not attachments:
        return "No Bill"
    for attachment in attachments:
        try:
            extraction = attachment.extraction
        except ObjectDoesNotExist:
            extraction = None
        if extraction is not None:
            return extraction.status
    return "OCR_PENDING"


def _expense_validation_status(expense_item: ExpenseItem) -> str:
    validation = getattr(expense_item, "validation", None)
    return validation.status if validation else "PENDING"


def _report_note_for_claim(claim: ReimbursementClaim) -> str:
    notes = [expense.review_notes for expense in claim.expense_items.all() if expense.review_notes]
    return "; ".join(notes[:3]) if notes else "-"


def _save_workbook_report(
    *,
    workbook: Workbook,
    report_type: str,
    filename: str,
    batch: MonthlyReimbursementBatch | None = None,
    claim: ReimbursementClaim | None = None,
    notes: str = "",
) -> GeneratedReport:
    buffer = BytesIO()
    workbook.save(buffer)
    report = GeneratedReport(
        batch=batch,
        claim=claim,
        report_type=report_type,
        original_filename=filename,
        notes=notes,
    )
    report.file.save(filename, ContentFile(buffer.getvalue()), save=False)
    report.save()
    return report


def _save_pdf_report(
    *,
    story: list,
    report_type: str,
    filename: str,
    batch: MonthlyReimbursementBatch | None = None,
    claim: ReimbursementClaim | None = None,
    notes: str = "",
) -> GeneratedReport:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    doc.build(story)
    report = GeneratedReport(
        batch=batch,
        claim=claim,
        report_type=report_type,
        original_filename=filename,
        notes=notes,
    )
    report.file.save(filename, ContentFile(buffer.getvalue()), save=False)
    report.save()
    return report


def generate_combined_excel_report(batch_id: int) -> GeneratedReport:
    batch = (
        MonthlyReimbursementBatch.objects.prefetch_related(
            "claims__employee",
            "claims__expense_items__attachments",
        )
        .filter(pk=batch_id)
        .first()
    )
    if not batch:
        raise ValueError("Reimbursement batch not found.")

    claims = list(batch.claims.select_related("employee").all())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Combined Report"

    rows = [
        ["Combined Reimbursement Report"],
        [f"Batch: {batch.title or f'Batch {batch.month}/{batch.year}'}"],
        [f"Generated At: {_generated_timestamp()}"],
        [],
        ["Summary"],
        ["Total Employees", batch.total_employees],
        ["Total Claims", len(claims)],
        ["Total Claimed Amount", _money(batch.total_claimed_amount)],
        ["Total Approved Amount", _money(batch.total_approved_amount)],
        [],
        [
            "Employee Code",
            "Employee Name",
            "Department",
            "Claim Status",
            "Expense Count",
            "Bills Attached Count",
            "Claimed Amount",
            "Approved Amount",
            "Review Notes",
        ],
    ]

    for row in rows:
        sheet.append(row)

    header_row = sheet.max_row
    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)

    for claim in claims:
        expense_items = list(claim.expense_items.prefetch_related("attachments").all())
        bills_attached_count = sum(1 for expense in expense_items if expense.attachments.exists())
        sheet.append(
            [
                claim.employee.employee_id,
                claim.employee.full_name,
                (claim.employee.structure_location.name if claim.employee.structure_location else "-"),
                claim.status,
                len(expense_items),
                bills_attached_count,
                _money(claim.total_claimed_amount),
                _money(claim.total_approved_amount),
                _report_note_for_claim(claim),
            ]
        )

    for column in ("A", "B", "C", "D", "E", "F", "G", "H", "I"):
        sheet.column_dimensions[column].width = 22

    filename = f"combined_reimbursement_report_batch_{batch.id}.xlsx"
    return _save_workbook_report(
        workbook=workbook,
        report_type=GeneratedReport.ReportType.COMBINED_EXCEL,
        filename=filename,
        batch=batch,
        notes="Combined monthly reimbursement Excel report.",
    )


def generate_employee_excel_report(claim_id: int) -> GeneratedReport:
    claim = (
        ReimbursementClaim.objects.select_related("employee", "batch")
        .prefetch_related("expense_items__attachments__extraction", "expense_items__validation")
        .filter(pk=claim_id)
        .first()
    )
    if not claim:
        raise ValueError("Reimbursement claim not found.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employee Report"

    rows = [
        ["Employee Reimbursement Claim Report"],
        [f"Employee: {claim.employee.full_name} ({claim.employee.employee_id})"],
        [f"Department: {claim.employee.structure_location.name if claim.employee.structure_location else '-'}",],
        [f"Batch: {claim.batch.title or f'Batch {claim.batch.month}/{claim.batch.year}'}"],
        [f"Claim Status: {claim.status}"],
        [f"Generated At: {_generated_timestamp()}"],
        ["Total Claimed Amount", _money(claim.total_claimed_amount)],
        ["Total Approved Amount", _money(claim.total_approved_amount)],
        [],
        [
            "Date",
            "Purpose",
            "Category",
            "Claimed Amount",
            "Approved Amount",
            "Bill Status",
            "OCR Status",
            "Validation Status",
            "Review Status",
            "Review Notes",
        ],
    ]

    for row in rows:
        sheet.append(row)

    header_row = sheet.max_row
    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)

    for expense_item in claim.expense_items.select_related("validation").prefetch_related("attachments__extraction").all():
        sheet.append(
            [
                expense_item.expense_date.isoformat() if expense_item.expense_date else "-",
                expense_item.description or expense_item.vendor_name or "-",
                expense_item.category,
                _money(expense_item.claimed_amount),
                _money(expense_item.approved_amount),
                _expense_bill_status(expense_item),
                _expense_ocr_status(expense_item),
                _expense_validation_status(expense_item),
                expense_item.status,
                expense_item.review_notes or "-",
            ]
        )

    for column in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J"):
        sheet.column_dimensions[column].width = 22

    filename = f"employee_reimbursement_claim_{claim.id}.xlsx"
    return _save_workbook_report(
        workbook=workbook,
        report_type=GeneratedReport.ReportType.EMPLOYEE_EXCEL,
        filename=filename,
        claim=claim,
        batch=claim.batch,
        notes="Employee reimbursement Excel report.",
    )


def generate_combined_pdf_report(batch_id: int) -> GeneratedReport:
    batch = (
        MonthlyReimbursementBatch.objects.prefetch_related("claims__employee", "claims__expense_items")
        .filter(pk=batch_id)
        .first()
    )
    if not batch:
        raise ValueError("Reimbursement batch not found.")

    styles = getSampleStyleSheet()
    story = [
        Paragraph("Combined Reimbursement Report", styles["Title"]),
        Spacer(1, 8),
        Paragraph(f"Batch: {batch.title or f'Batch {batch.month}/{batch.year}'}", styles["Normal"]),
        Paragraph(f"Generated At: {_generated_timestamp()}", styles["Normal"]),
        Spacer(1, 10),
        Paragraph(
            f"Total Employees: {batch.total_employees} | Total Claims: {batch.claims.count()} | "
            f"Total Claimed: {_money(batch.total_claimed_amount)} | Total Approved: {_money(batch.total_approved_amount)}",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]

    table_data = [[
        "Employee Code",
        "Employee Name",
        "Department",
        "Claim Status",
        "Expense Count",
        "Claimed Amount",
        "Approved Amount",
    ]]
    for claim in batch.claims.select_related("employee").all():
        table_data.append([
            claim.employee.employee_id,
            claim.employee.full_name,
           (claim.employee.structure_location.name if claim.employee.structure_location else "-"),
            claim.status,
            str(claim.expense_items.count()),
            _money(claim.total_claimed_amount),
            _money(claim.total_approved_amount),
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dfeef8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(
        Paragraph(
            f"Grand Total Claimed: {_money(batch.total_claimed_amount)} | Grand Total Approved: {_money(batch.total_approved_amount)}",
            styles["Heading3"],
        )
    )

    filename = f"combined_reimbursement_report_batch_{batch.id}.pdf"
    return _save_pdf_report(
        story=story,
        report_type=GeneratedReport.ReportType.COMBINED_PDF,
        filename=filename,
        batch=batch,
        notes="Combined monthly reimbursement PDF report.",
    )


def generate_employee_pdf_report(claim_id: int) -> GeneratedReport:
    claim = (
        ReimbursementClaim.objects.select_related("employee", "batch")
        .prefetch_related("expense_items__attachments__extraction", "expense_items__validation")
        .filter(pk=claim_id)
        .first()
    )
    if not claim:
        raise ValueError("Reimbursement claim not found.")

    styles = getSampleStyleSheet()
    story = [
        Paragraph("Employee Reimbursement Claim Report", styles["Title"]),
        Spacer(1, 8),
        Paragraph(
            f"Employee: {claim.employee.full_name} ({claim.employee.employee_id}) | Department: {(claim.employee.structure_location.name if claim.employee.structure_location else '-')}",
            styles["Normal"],
        ),
        Paragraph(
            f"Batch: {claim.batch.title or f'Batch {claim.batch.month}/{claim.batch.year}'} | Claim Status: {claim.status}",
            styles["Normal"],
        ),
        Paragraph(f"Generated At: {_generated_timestamp()}", styles["Normal"]),
        Spacer(1, 10),
        Paragraph(
            f"Total Claimed: {_money(claim.total_claimed_amount)} | Total Approved: {_money(claim.total_approved_amount)}",
            styles["Heading3"],
        ),
        Spacer(1, 12),
    ]

    table_data = [[
        "Date",
        "Purpose",
        "Category",
        "Claimed",
        "Approved",
        "Bill",
        "OCR",
        "Validation",
        "Review",
    ]]
    for expense_item in claim.expense_items.select_related("validation").prefetch_related("attachments__extraction").all():
        table_data.append([
            expense_item.expense_date.isoformat() if expense_item.expense_date else "-",
            expense_item.description or expense_item.vendor_name or "-",
            expense_item.category,
            _money(expense_item.claimed_amount),
            _money(expense_item.approved_amount),
            _expense_bill_status(expense_item),
            _expense_ocr_status(expense_item),
            _expense_validation_status(expense_item),
            expense_item.status,
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dfeef8")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(table)

    filename = f"employee_reimbursement_claim_{claim.id}.pdf"
    return _save_pdf_report(
        story=story,
        report_type=GeneratedReport.ReportType.EMPLOYEE_PDF,
        filename=filename,
        claim=claim,
        batch=claim.batch,
        notes="Employee reimbursement PDF report.",
    )


def _quick_claim_source_label(classification_source: str) -> str:
    return {
        ExpenseItem.ClassificationSource.VENDOR_RULE: "RULE",
        ExpenseItem.ClassificationSource.TEXT_HEURISTIC: "OCR",
        ExpenseItem.ClassificationSource.MANUAL: "MANUAL",
        ExpenseItem.ClassificationSource.LLM_FALLBACK: "LLM",
    }.get(classification_source, "OCR")


def _quick_claim_review_status(draft) -> str:
    if draft.requires_manual_review:
        return "Needs Review"
    if draft.manually_reviewed:
        return "Manually Verified"
    return "Verified"


def _quick_claim_upload(upload_id: int) -> SmartReimbursementUpload:
    upload = (
        SmartReimbursementUpload.objects.select_related(
            "claim__employee",
            "claim__batch",
            "created_by_employee",
            "excel_report",
            "pdf_report",
        )
        .prefetch_related("draft_expenses__bill_file")
        .filter(pk=upload_id)
        .first()
    )
    if not upload or not upload.claim_id:
        raise ValueError("Confirmed Smart Reimbursement Upload not found.")
    return upload


def generate_quick_claim_excel_report(upload_id: int) -> GeneratedReport:
    upload = _quick_claim_upload(upload_id)
    if upload.excel_report_id and upload.excel_report.report_type == GeneratedReport.ReportType.QUICK_CLAIM_EXCEL:
        return upload.excel_report

    claim = upload.claim
    employee = claim.employee
    system_setting = get_system_settings()
    drafts = list(upload.draft_expenses.select_related("bill_file").order_by("id"))
    manual_count = sum(draft.manually_reviewed for draft in drafts)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quick Claim Summary"
    sheet.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="16324F")
    section_fill = PatternFill("solid", fgColor="DCECF7")
    header_fill = PatternFill("solid", fgColor="0F8EC8")
    total_fill = PatternFill("solid", fgColor="EAF7F2")
    thin_border = Border(
        left=Side(style="thin", color="D4E0EA"),
        right=Side(style="thin", color="D4E0EA"),
        top=Side(style="thin", color="D4E0EA"),
        bottom=Side(style="thin", color="D4E0EA"),
    )

    sheet.merge_cells("A1:K1")
    title_cell = sheet["A1"]
    title_cell.value = "Quick Claim Reimbursement Report"
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 32

    metadata = [
        ("Company", system_setting.company_name),
        ("Employee", employee.full_name),
        ("Department",(employee.structure_location.name if employee.structure_location else '') or upload.employee_department or "-"),
        ("Claim Month / Year", f"{month_name[upload.month]} {upload.year}"),
        ("Generated", _generated_timestamp()),
        ("Finance Recipient", upload.recipient_email or "-"),
        ("Claim Source", "QUICK_CLAIM / SMART_UPLOAD"),
        ("Uploaded / Processed / Failed", f"{upload.total_files} / {upload.processed_files} / {upload.failed_files}"),
        ("Manually Edited Rows", manual_count),
        ("Final Confirmed Amount", Decimal(claim.total_claimed_amount)),
    ]
    for row_index, (label, value) in enumerate(metadata, start=3):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=5)
        sheet.cell(row=row_index, column=1).fill = section_fill
        sheet.cell(row=row_index, column=1).font = Font(bold=True, color="16324F")
        sheet.cell(row=row_index, column=1).border = thin_border
        for column in range(2, 6):
            sheet.cell(row=row_index, column=column).border = thin_border
        if label == "Final Confirmed Amount":
            sheet.cell(row=row_index, column=2).number_format = '#,##0.00'

    header_row = 14
    headers = [
        "S.No",
        "Bill File Name",
        "Bill Date",
        "Vendor",
        "Purpose / Details of Claim",
        "Category",
        "Amount",
        "Classification Confidence",
        "Source",
        "Review Status",
        "Remarks",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    sheet.row_dimensions[header_row].height = 34
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:K{header_row + len(drafts)}"

    for serial_number, draft in enumerate(drafts, start=1):
        row = header_row + serial_number
        values = [
            serial_number,
            draft.bill_file.original_filename,
            draft.expense_date,
            draft.vendor_name,
            draft.purpose,
            draft.get_category_display(),
            draft.amount,
            draft.category_confidence,
            _quick_claim_source_label(draft.classification_source),
            _quick_claim_review_status(draft),
            draft.remarks or "-",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=column in {2, 4, 5, 11})
        sheet.cell(row=row, column=3).number_format = "dd-mmm-yyyy"
        sheet.cell(row=row, column=7).number_format = '#,##0.00'
        sheet.cell(row=row, column=8).number_format = "0%"

    total_row = header_row + len(drafts) + 1
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    sheet.cell(row=total_row, column=1, value="Final Confirmed Total")
    sheet.cell(row=total_row, column=7, value=Decimal(claim.total_claimed_amount))
    for column in range(1, 12):
        cell = sheet.cell(row=total_row, column=column)
        cell.fill = total_fill
        cell.font = Font(bold=True, color="0A6A4D")
        cell.border = thin_border
    sheet.cell(row=total_row, column=7).number_format = '#,##0.00'

    widths = {1: 8, 2: 28, 3: 15, 4: 22, 5: 40, 6: 18, 7: 15, 8: 18, 9: 12, 10: 20, 11: 32}
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = min(width, 45)

    filename = f"quick_claim_{upload.id}_{employee.employee_id}.xlsx"
    return _save_workbook_report(
        workbook=workbook,
        report_type=GeneratedReport.ReportType.QUICK_CLAIM_EXCEL,
        filename=filename,
        batch=claim.batch,
        claim=claim,
        notes=f"Quick Claim Excel report for Smart Upload #{upload.id}.",
    )


def generate_quick_claim_pdf_report(upload_id: int) -> GeneratedReport:
    upload = _quick_claim_upload(upload_id)
    if upload.pdf_report_id and upload.pdf_report.report_type == GeneratedReport.ReportType.QUICK_CLAIM_PDF:
        return upload.pdf_report

    claim = upload.claim
    employee = claim.employee
    system_setting = get_system_settings()
    drafts = list(upload.draft_expenses.select_related("bill_file").order_by("id"))
    manual_count = sum(draft.manually_reviewed for draft in drafts)
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle(
        "QuickClaimBody",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#24384D"),
    )
    heading_style = ParagraphStyle(
        "QuickClaimHeading",
        parent=styles["Title"],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#16324F"),
        spaceAfter=8,
    )

    story = [
        Paragraph("Quick Claim Reimbursement Report", heading_style),
        Paragraph(escape(system_setting.company_name), styles["Heading3"]),
        Spacer(1, 8),
        Paragraph(
            escape(
                f"Employee: {employee.full_name} | Department: {(employee.structure_location.name if employee.structure_location else '') or upload.employee_department or '-'} | "
                f"Claim Month: {month_name[upload.month]} {upload.year}"
            ),
            body_style,
        ),
        Paragraph(f"Generated: {escape(_generated_timestamp())}", body_style),
        Paragraph(f"Finance Recipient: {escape(upload.recipient_email or '-')}", body_style),
        Paragraph(
            escape(
                f"Final Confirmed Amount: INR {_money(claim.total_claimed_amount)} | "
                f"Uploaded: {upload.total_files} | Processed: {upload.processed_files} | "
                f"Failed: {upload.failed_files} | Manual Reviews: {manual_count}"
            ),
            styles["Heading3"],
        ),
        Spacer(1, 12),
    ]

    table_data = [["Date", "Vendor", "Purpose", "Category", "Amount", "Status"]]
    for draft in drafts:
        table_data.append(
            [
                draft.expense_date.strftime("%d-%b-%Y") if draft.expense_date else "-",
                Paragraph(escape(draft.vendor_name or "-"), body_style),
                Paragraph(escape(draft.purpose or "-"), body_style),
                draft.get_category_display(),
                f"INR {_money(draft.amount)}",
                _quick_claim_review_status(draft),
            ]
        )
    table_data.append(["", "", "", "Final Total", f"INR {_money(claim.total_claimed_amount)}", "Confirmed"])

    table = Table(table_data, repeatRows=1, colWidths=[68, 110, 260, 90, 85, 100])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F8EC8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEADING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C6D5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F4F8FB")]),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EAF7F2")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(table)

    filename = f"quick_claim_{upload.id}_{employee.employee_id}.pdf"
    return _save_pdf_report(
        story=story,
        report_type=GeneratedReport.ReportType.QUICK_CLAIM_PDF,
        filename=filename,
        batch=claim.batch,
        claim=claim,
        notes=f"Quick Claim PDF report for Smart Upload #{upload.id}.",
    )

