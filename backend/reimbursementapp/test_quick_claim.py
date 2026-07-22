import shutil
import tempfile
import base64
import smtplib
import zipfile
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import fitz
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework.test import APITestCase

from .models import (
    BillExtraction,
    DraftExtractedExpense,
    EmployeeReimbursementProfile,
    EmailDispatchLog,
    ExpenseAttachment,
    ExpenseItem,
    GeneratedReport,
    ReimbursementClaim,
    SmartReimbursementUpload,
    SmartUploadedBillFile,
    SystemSetting,
    UploadedReimbursementForm,
)
from .quick_claim_services import confirm_and_send_upload, recipient_is_allowed, validate_recipients
from .tasks import enqueue_smart_upload, process_bill_file


TEST_MEDIA_ROOT = tempfile.mkdtemp()
PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class QuickClaimApiTests(APITestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(username="employee", password="secret")
        self.other_user = user_model.objects.create_user(username="other", password="secret")
        self.finance_user = user_model.objects.create_user(
            username="finance",
            password="secret",
            is_staff=True,
        )
        self.profile = EmployeeReimbursementProfile.objects.create(
            user=self.user,
            employee_name="Jithin Raj",
            department="Technology",
            default_claim_month=6,
            default_claim_year=2026,
            finance_head_email="finance@vbsai.com",
            cc_emails=[],
            is_complete=True,
        )

    def _create_upload(self, **overrides):
        defaults = {
            "created_by": self.user,
            "employee_name": "Jithin Raj",
            "month": 6,
            "year": 2026,
            "total_files": 1,
            "status": SmartReimbursementUpload.Status.NEEDS_REVIEW,
        }
        defaults.update(overrides)
        return SmartReimbursementUpload.objects.create(**defaults)

    def _zip_file(self, entries):
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w") as archive:
            for filename, content in entries:
                archive.writestr(filename, content)
        return SimpleUploadedFile("monthly-bills.zip", buffer.getvalue(), content_type="application/zip")

    def _ocr_success(self, attachment):
        filename = (attachment.original_filename or "").lower()
        vendor = "Swiggy" if "swiggy" in filename else "Rapido"
        amount = Decimal("599.00") if vendor == "Swiggy" else Decimal("245.00")
        return BillExtraction.objects.create(
            attachment=attachment,
            status=BillExtraction.Status.COMPLETED,
            raw_text=f"{vendor} invoice total {amount}",
            extracted_vendor=vendor,
            extracted_date=date(2026, 6, 24),
            extracted_amount=amount,
        )

    def _create_confirmable_upload(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])
        upload = self._create_upload(
            status=SmartReimbursementUpload.Status.READY_TO_CONFIRM,
            recipient_email="finance@vbsai.com",
            processed_files=1,
        )
        bill = SmartUploadedBillFile.objects.create(
            upload=upload,
            file=SimpleUploadedFile("bill.png", PNG_BYTES),
            original_filename="bill.png",
            detected_mime_type="image/png",
            file_size=len(PNG_BYTES),
            content_sha256="9" * 64,
            status=SmartUploadedBillFile.Status.PROCESSED,
        )
        DraftExtractedExpense.objects.create(
            upload=upload,
            bill_file=bill,
            expense_date=date(2026, 6, 24),
            vendor_name="Rapido",
            purpose="Travel to client meeting",
            category=ExpenseItem.Category.TRAVEL,
            amount=Decimal("245.00"),
            requires_manual_review=False,
        )
        self.client.force_authenticate(self.user)
        return upload

    def test_upload_requires_authentication(self):
        response = self.client.post(
            reverse("quick-claim-upload"),
            {"month": 6, "year": 2026, "files[]": SimpleUploadedFile("bill.png", PNG_BYTES)},
            format="multipart",
        )

        self.assertIn(response.status_code, {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN})

    @patch("reimbursements.views.enqueue_smart_upload")
    def test_upload_validates_content_and_creates_staging_rows(self, mocked_enqueue):
        self.client.force_authenticate(self.user)

        response = self.client.post(
            reverse("quick-claim-upload"),
            {
                "employee_name": "Jithin Raj",
                "month": 6,
                "year": 2026,
                "files[]": SimpleUploadedFile("bill.png", PNG_BYTES, content_type="text/plain"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_202_ACCEPTED)
        upload = SmartReimbursementUpload.objects.get()
        bill = upload.bill_files.get()
        self.assertEqual(bill.detected_mime_type, "image/png")
        self.assertEqual(upload.created_by, self.user)
        self.assertEqual(upload.reimbursement_profile, self.profile)
        self.assertEqual(upload.employee_name, "Jithin Raj")
        mocked_enqueue.assert_called_once_with(upload.id)

    @patch("reimbursements.views.enqueue_smart_upload")
    def test_upload_blocks_authenticated_employee_with_incomplete_profile(self, mocked_enqueue):
        self.profile.is_complete = False
        self.profile.save(update_fields=["is_complete", "updated_at"])
        self.client.force_authenticate(self.user)

        response = self.client.post(
            reverse("quick-claim-upload"),
            {
                "month": 6,
                "year": 2026,
                "files[]": SimpleUploadedFile("bill.png", PNG_BYTES, content_type="image/png"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("completed reimbursement profile", str(response.data))
        self.assertFalse(SmartReimbursementUpload.objects.exists())
        mocked_enqueue.assert_not_called()

    @patch("reimbursements.views.enqueue_smart_upload")
    def test_upload_rejects_disallowed_actual_mime_type(self, mocked_enqueue):
        self.client.force_authenticate(self.user)

        response = self.client.post(
            reverse("quick-claim-upload"),
            {
                "month": 6,
                "year": 2026,
                "files[]": SimpleUploadedFile("fake.pdf", b"plain text executable content"),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(SmartReimbursementUpload.objects.exists())
        mocked_enqueue.assert_not_called()

    @patch("reimbursements.tasks.run_ocr_for_attachment")
    def test_zip_upload_processes_valid_bills_and_records_unsupported_entries(self, mocked_ocr):
        mocked_ocr.side_effect = self._ocr_success
        self.client.force_authenticate(self.user)

        response = self.client.post(
            reverse("quick-claim-upload"),
            {
                "employee_name": "Jithin Raj",
                "month": 6,
                "year": 2026,
                "files[]": self._zip_file(
                    [
                        ("rapido-trip.png", PNG_BYTES),
                        ("notes.txt", b"not a reimbursement bill"),
                    ]
                ),
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_202_ACCEPTED)
        upload = SmartReimbursementUpload.objects.get()
        self.assertEqual(upload.total_files, 2)
        self.assertEqual(upload.processed_files, 1)
        self.assertEqual(upload.failed_files, 1)
        self.assertEqual(upload.status, SmartReimbursementUpload.Status.NEEDS_REVIEW)
        self.assertEqual(upload.bill_files.count(), 2)
        self.assertEqual(upload.bill_files.filter(status=SmartUploadedBillFile.Status.PROCESSED).count(), 1)
        failed = upload.bill_files.get(status=SmartUploadedBillFile.Status.FAILED)
        self.assertEqual(failed.original_filename, "notes.txt")
        self.assertIn("not allowed", failed.error_message)
        self.assertEqual(upload.draft_expenses.count(), 1)

    @patch("reimbursements.tasks.run_ocr_for_attachment")
    def test_multi_file_upload_processes_supported_files_and_records_unsupported_companions(self, mocked_ocr):
        mocked_ocr.side_effect = self._ocr_success
        self.client.force_authenticate(self.user)

        response = self.client.post(
            reverse("quick-claim-upload"),
            {
                "employee_name": "Jithin Raj",
                "month": 6,
                "year": 2026,
                "files[]": [
                    SimpleUploadedFile("rapido-trip.png", PNG_BYTES, content_type="image/png"),
                    SimpleUploadedFile("swiggy-meal.png", PNG_BYTES, content_type="image/png"),
                    SimpleUploadedFile("readme.txt", b"ignore me", content_type="text/plain"),
                ],
            },
            format="multipart",
        )

        self.assertEqual(response.status_code, status.HTTP_202_ACCEPTED)
        upload = SmartReimbursementUpload.objects.get()
        self.assertEqual(upload.total_files, 3)
        self.assertEqual(upload.processed_files, 2)
        self.assertEqual(upload.failed_files, 1)
        self.assertEqual(upload.status, SmartReimbursementUpload.Status.NEEDS_REVIEW)
        self.assertEqual(upload.draft_expenses.count(), 2)
        self.assertEqual(
            set(upload.draft_expenses.values_list("vendor_name", flat=True)),
            {"Rapido", "Swiggy"},
        )

    def test_status_is_scoped_to_authenticated_owner(self):
        upload = self._create_upload()
        self.client.force_authenticate(self.other_user)

        response = self.client.get(reverse("quick-claim-status", args=[upload.id]))

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_finance_admin_can_access_employee_upload_status(self):
        upload = self._create_upload()
        self.client.force_authenticate(self.finance_user)

        response = self.client.get(reverse("quick-claim-status", args=[upload.id]))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["id"], upload.id)

    def test_employee_cannot_access_full_app_restricted_apis(self):
        self.client.force_authenticate(self.user)

        responses = [
            self.client.get(reverse("generated-report-list")),
            self.client.get(reverse("system-settings")),
            self.client.get(reverse("finance-review-summary")),
            self.client.get(reverse("email-log-list")),
        ]

        for response in responses:
            self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
            self.assertEqual(response.data["detail"], "You do not have permission to access this page.")

    def test_finance_admin_can_access_full_app_restricted_apis(self):
        self.client.force_authenticate(self.finance_user)

        responses = [
            self.client.get(reverse("generated-report-list")),
            self.client.get(reverse("system-settings")),
            self.client.get(reverse("finance-review-summary")),
            self.client.get(reverse("email-log-list")),
        ]

        for response in responses:
            self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_inline_edit_marks_draft_as_manually_reviewed(self):
        upload = self._create_upload()
        bill = SmartUploadedBillFile.objects.create(
            upload=upload,
            file=SimpleUploadedFile("bill.png", PNG_BYTES),
            original_filename="bill.png",
            detected_mime_type="image/png",
            file_size=len(PNG_BYTES),
            content_sha256="a" * 64,
            status=SmartUploadedBillFile.Status.PROCESSED,
        )
        draft = DraftExtractedExpense.objects.create(
            upload=upload,
            bill_file=bill,
            category=ExpenseItem.Category.OTHERS,
            amount=Decimal("120.00"),
            requires_manual_review=True,
        )
        self.client.force_authenticate(self.user)

        response = self.client.patch(
            reverse("quick-claim-update-item", args=[upload.id, draft.id]),
            {
                "expense_date": "2026-06-24",
                "vendor_name": "Rapido",
                "purpose": "Travel - Rapido",
                "category": "TRAVEL",
                "amount": "245.00",
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        draft.refresh_from_db()
        self.assertTrue(draft.manually_reviewed)
        self.assertFalse(draft.requires_manual_review)
        self.assertEqual(draft.classification_source, ExpenseItem.ClassificationSource.MANUAL)
        self.assertEqual(draft.reviewed_by, self.user)

    def test_inline_edit_keeps_incomplete_rows_flagged_for_review(self):
        upload = self._create_upload()
        bill = SmartUploadedBillFile.objects.create(
            upload=upload,
            file=SimpleUploadedFile("bill.png", PNG_BYTES),
            original_filename="bill.png",
            detected_mime_type="image/png",
            file_size=len(PNG_BYTES),
            content_sha256="d" * 64,
            status=SmartUploadedBillFile.Status.PROCESSED,
        )
        draft = DraftExtractedExpense.objects.create(
            upload=upload,
            bill_file=bill,
            expense_date=date(2026, 6, 24),
            vendor_name="Rapido",
            purpose="Travel",
            category=ExpenseItem.Category.TRAVEL,
            amount=Decimal("120.00"),
        )
        self.client.force_authenticate(self.user)

        response = self.client.patch(
            reverse("quick-claim-update-item", args=[upload.id, draft.id]),
            {"purpose": ""},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        draft.refresh_from_db()
        self.assertTrue(draft.manually_reviewed)
        self.assertTrue(draft.requires_manual_review)
        self.assertEqual(draft.classification_source, ExpenseItem.ClassificationSource.MANUAL)
        self.assertEqual(draft.category_confidence, 1.0)

    def test_confirm_and_send_rejects_non_allowlisted_recipient(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])
        self.profile.finance_head_email = "finance@external.example"
        self.profile.save(update_fields=["finance_head_email", "updated_at"])
        upload = self._create_upload(recipient_email="finance@external.example")
        self.client.force_authenticate(self.user)

        response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    @patch("reimbursements.email_services.EmailMessage")
    def test_confirm_and_send_api_blocks_duplicate_send_without_duplicate_records(self, mocked_email_class):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])
        mocked_email_class.return_value.send.return_value = 1
        upload = self._create_upload(
            status=SmartReimbursementUpload.Status.READY_TO_CONFIRM,
            recipient_email="finance@vbsai.com",
            processed_files=1,
        )
        bill = SmartUploadedBillFile.objects.create(
            upload=upload,
            file=SimpleUploadedFile("bill.png", PNG_BYTES),
            original_filename="bill.png",
            detected_mime_type="image/png",
            file_size=len(PNG_BYTES),
            content_sha256="e" * 64,
            status=SmartUploadedBillFile.Status.PROCESSED,
        )
        DraftExtractedExpense.objects.create(
            upload=upload,
            bill_file=bill,
            expense_date=date(2026, 6, 24),
            vendor_name="Rapido",
            purpose="Travel to client meeting",
            category=ExpenseItem.Category.TRAVEL,
            amount=Decimal("245.00"),
            requires_manual_review=False,
        )
        self.client.force_authenticate(self.user)

        first_response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {"cc_emails": ["manager@vbsai.com"]},
            format="json",
        )
        second_response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )

        self.assertEqual(first_response.status_code, status.HTTP_202_ACCEPTED)
        self.assertEqual(second_response.status_code, status.HTTP_409_CONFLICT)
        upload.refresh_from_db()
        self.assertEqual(upload.status, SmartReimbursementUpload.Status.SENT)
        self.assertEqual(upload.claim.expense_items.count(), 1)
        self.assertEqual(EmailDispatchLog.objects.filter(smart_reimbursement_upload=upload).count(), 1)
        self.assertEqual(mocked_email_class.return_value.send.call_count, 1)

    @patch("reimbursements.email_services.EmailMessage")
    def test_confirm_and_send_api_blocks_unreviewed_drafts_before_smtp(self, mocked_email_class):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])
        upload = self._create_upload(
            status=SmartReimbursementUpload.Status.NEEDS_REVIEW,
            recipient_email="finance@vbsai.com",
            processed_files=1,
        )
        bill = SmartUploadedBillFile.objects.create(
            upload=upload,
            file=SimpleUploadedFile("bill.png", PNG_BYTES),
            original_filename="bill.png",
            detected_mime_type="image/png",
            file_size=len(PNG_BYTES),
            content_sha256="f" * 64,
            status=SmartUploadedBillFile.Status.PROCESSED,
        )
        draft = DraftExtractedExpense.objects.create(
            upload=upload,
            bill_file=bill,
            expense_date=date(2026, 6, 24),
            vendor_name="",
            purpose="Travel",
            category=ExpenseItem.Category.TRAVEL,
            amount=Decimal("245.00"),
            requires_manual_review=True,
        )
        self.client.force_authenticate(self.user)

        blocked_response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )
        draft.vendor_name = "Rapido"
        draft.requires_manual_review = False
        draft.manually_reviewed = True
        draft.classification_source = ExpenseItem.ClassificationSource.MANUAL
        draft.category_confidence = 1.0
        draft.save()
        mocked_email_class.return_value.send.return_value = 1
        success_response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )

        self.assertEqual(blocked_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("items", blocked_response.data)
        self.assertEqual(success_response.status_code, status.HTTP_202_ACCEPTED)
        mocked_email_class.return_value.send.assert_called_once_with(fail_silently=False)

    def test_quick_claim_api_contract_auth_owner_validation_and_cancel(self):
        upload = self._create_upload()

        unauthenticated = self.client.get(reverse("quick-claim-status", args=[upload.id]))
        self.client.force_authenticate(self.other_user)
        wrong_owner = self.client.get(reverse("quick-claim-draft-expenses", args=[upload.id]))
        self.client.force_authenticate(self.user)
        missing_files = self.client.post(
            reverse("quick-claim-upload"),
            {"month": 6, "year": 2026},
            format="multipart",
        )
        cancel_response = self.client.post(reverse("quick-claim-cancel", args=[upload.id]), {}, format="json")
        upload.refresh_from_db()

        self.assertIn(unauthenticated.status_code, {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN})
        self.assertEqual(wrong_owner.status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(missing_files.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(cancel_response.status_code, status.HTTP_200_OK)
        self.assertEqual(upload.status, SmartReimbursementUpload.Status.CANCELLED)

    def test_existing_excel_form_upload_endpoint_requires_finance_access(self):
        response = self.client.post(
            reverse("upload-reimbursement-form"),
            {"file": SimpleUploadedFile("not-a-workbook.txt", b"not excel")},
            format="multipart",
        )

        self.assertIn(response.status_code, {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN})
        self.assertFalse(UploadedReimbursementForm.objects.exists())

        self.client.force_authenticate(self.finance_user)
        finance_response = self.client.post(
            reverse("upload-reimbursement-form"),
            {"file": SimpleUploadedFile("not-a-workbook.txt", b"not excel")},
            format="multipart",
        )

        self.assertEqual(finance_response.status_code, status.HTTP_201_CREATED)
        self.assertTrue(UploadedReimbursementForm.objects.exists())

    def test_confirm_and_send_requires_authentication(self):
        upload = self._create_upload(recipient_email="finance@vbsai.com")

        response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )

        self.assertIn(response.status_code, {status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN})

    @patch("reimbursements.views.confirm_and_send_quick_claim.delay")
    def test_confirm_and_send_queues_authenticated_owned_upload(self, mocked_delay):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])
        upload = self._create_upload(
            status=SmartReimbursementUpload.Status.READY_TO_CONFIRM,
            recipient_email="finance@vbsai.com",
        )
        bill = SmartUploadedBillFile.objects.create(
            upload=upload,
            file=SimpleUploadedFile("bill.png", PNG_BYTES),
            original_filename="bill.png",
            detected_mime_type="image/png",
            file_size=len(PNG_BYTES),
            content_sha256="g" * 64,
            status=SmartUploadedBillFile.Status.PROCESSED,
        )
        DraftExtractedExpense.objects.create(
            upload=upload,
            bill_file=bill,
            expense_date=date(2026, 6, 24),
            vendor_name="Rapido",
            purpose="Travel to client meeting",
            category=ExpenseItem.Category.TRAVEL,
            amount=Decimal("245.00"),
            requires_manual_review=False,
        )
        mocked_delay.return_value = SimpleNamespace(id="task-123")
        self.client.force_authenticate(self.user)

        response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_202_ACCEPTED)
        mocked_delay.assert_called_once_with(upload.id, self.user.id)
        upload.refresh_from_db()
        self.assertEqual(upload.status, SmartReimbursementUpload.Status.CONFIRMING)

    @patch(
        "reimbursements.quick_claim_services.generate_quick_claim_excel_report",
        side_effect=RuntimeError("Workbook could not be saved."),
    )
    def test_report_generation_failure_returns_json_and_resets_status(self, mocked_report):
        upload = self._create_confirmable_upload()

        response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertEqual(response.data["stage"], "report")
        self.assertIn("Report generation failed", response.data["detail"])
        upload.refresh_from_db()
        self.assertEqual(upload.status, SmartReimbursementUpload.Status.READY_TO_CONFIRM)
        self.assertIn("Workbook could not be saved", upload.error_message)
        self.assertFalse(EmailDispatchLog.objects.filter(smart_reimbursement_upload=upload).exists())
        mocked_report.assert_called_once_with(upload.id)

    @patch(
        "reimbursements.email_services.EmailMessage.attach_file",
        side_effect=OSError("Report attachment could not be opened."),
    )
    def test_email_attachment_failure_returns_json_and_keeps_reports(self, mocked_attach):
        upload = self._create_confirmable_upload()

        response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_502_BAD_GATEWAY)
        self.assertEqual(response.data["stage"], "email")
        self.assertIn("Report attachment could not be opened", response.data["detail"])
        upload.refresh_from_db()
        self.assertEqual(upload.status, SmartReimbursementUpload.Status.READY_TO_CONFIRM)
        self.assertIsNotNone(upload.excel_report_id)
        self.assertIsNotNone(upload.pdf_report_id)
        log = EmailDispatchLog.objects.get(smart_reimbursement_upload=upload)
        self.assertEqual(log.status, EmailDispatchLog.Status.FAILED)
        self.assertEqual(log.attached_reports.count(), 2)
        self.assertIn("Report attachment could not be opened", log.error_message)
        mocked_attach.assert_called()

    @patch(
        "reimbursements.email_services.EmailMessage.send",
        side_effect=smtplib.SMTPServerDisconnected("Connection unexpectedly closed"),
    )
    def test_smtp_failure_returns_json_and_records_failed_email_log(self, mocked_send):
        upload = self._create_confirmable_upload()

        response = self.client.post(
            reverse("quick-claim-confirm-and-send", args=[upload.id]),
            {},
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_502_BAD_GATEWAY)
        self.assertEqual(response.data["stage"], "email")
        self.assertEqual(response.data["detail"], "Connection unexpectedly closed")
        upload.refresh_from_db()
        self.assertEqual(upload.status, SmartReimbursementUpload.Status.READY_TO_CONFIRM)
        self.assertEqual(upload.error_message, "Connection unexpectedly closed")
        self.assertIsNotNone(upload.excel_report_id)
        self.assertIsNotNone(upload.pdf_report_id)
        log = EmailDispatchLog.objects.get(smart_reimbursement_upload=upload)
        self.assertEqual(log.status, EmailDispatchLog.Status.FAILED)
        self.assertEqual(log.error_message, "Connection unexpectedly closed")
        self.assertEqual(response.data["email_log_id"], log.id)
        mocked_send.assert_called_once_with(fail_silently=False)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT)
class QuickClaimTaskTests(TestCase):
    def setUp(self):
        Path(TEST_MEDIA_ROOT).mkdir(parents=True, exist_ok=True)
        self.user = get_user_model().objects.create_user(username="task-user")
        self.upload = SmartReimbursementUpload.objects.create(
            created_by=self.user,
            employee_name="Jithin Raj",
            month=6,
            year=2026,
            total_files=1,
            status=SmartReimbursementUpload.Status.QUEUED,
        )
        self.profile = EmployeeReimbursementProfile.objects.create(
            user=self.user,
            employee_name="Jithin Raj",
            department="Technology",
            default_claim_month=6,
            default_claim_year=2026,
            finance_head_email="finance@vbsai.com",
            cc_emails=[],
            is_complete=True,
        )
        self.bill = SmartUploadedBillFile.objects.create(
            upload=self.upload,
            file=SimpleUploadedFile("rapido.png", PNG_BYTES),
            original_filename="rapido.png",
            detected_mime_type="image/png",
            file_size=len(PNG_BYTES),
            content_sha256="b" * 64,
        )
        self.attachment = ExpenseAttachment.objects.create(
            file=SimpleUploadedFile("rapido.png", PNG_BYTES),
            original_filename="rapido.png",
            file_type="image/png",
        )
        self.bill.expense_attachment = self.attachment
        self.bill.save(update_fields=["expense_attachment", "updated_at"])

    def _prepare_confirmable_upload(self):
        SystemSetting.objects.create(
            company_name="NL Technologies Pvt. Ltd.",
            quick_claim_allowed_recipient_domains=["vbsai.com"],
        )
        self.profile.finance_head_email = "finance@vbsai.com"
        self.profile.cc_emails = ["manager@vbsai.com", "hr@vbsai.com"]
        self.profile.save(update_fields=["finance_head_email", "cc_emails", "updated_at"])
        self.upload.recipient_email = "finance@vbsai.com"
        self.upload.cc_emails = ["manager@vbsai.com", "hr@vbsai.com"]
        self.upload.status = SmartReimbursementUpload.Status.READY_TO_CONFIRM
        self.upload.processed_files = 1
        self.upload.save(
            update_fields=["recipient_email", "cc_emails", "status", "processed_files", "updated_at"]
        )
        return DraftExtractedExpense.objects.create(
            upload=self.upload,
            bill_file=self.bill,
            expense_date=date(2026, 6, 24),
            vendor_name="Rapido",
            purpose="Travel from office to client meeting",
            remarks="Verified against the attached receipt.",
            category=ExpenseItem.Category.TRAVEL,
            amount=Decimal("245.00"),
            classification_source=ExpenseItem.ClassificationSource.MANUAL,
            category_confidence=1.0,
            requires_manual_review=False,
            manually_reviewed=True,
        )

    @patch("reimbursements.tasks.run_ocr_for_attachment")
    def test_file_task_is_idempotent_after_success(self, mocked_ocr):
        def completed_extraction(attachment):
            return BillExtraction.objects.create(
                attachment=attachment,
                status=BillExtraction.Status.COMPLETED,
                raw_text="Rapido ride invoice total 245",
                extracted_vendor="Rapido",
                extracted_date=date(2026, 6, 24),
                extracted_amount=Decimal("245.00"),
            )

        mocked_ocr.side_effect = completed_extraction

        first_result = process_bill_file.apply(args=[self.bill.id]).get()
        second_result = process_bill_file.apply(args=[self.bill.id]).get()

        self.assertEqual(first_result["status"], "processed")
        self.assertTrue(second_result["idempotent"])
        self.assertEqual(DraftExtractedExpense.objects.filter(bill_file=self.bill).count(), 1)
        self.assertEqual(mocked_ocr.call_count, 1)
        self.upload.refresh_from_db()
        self.assertEqual(self.upload.processed_files, 1)

    @patch("reimbursements.tasks.run_ocr_for_attachment")
    def test_eager_chord_processes_file_and_runs_finalizer(self, mocked_ocr):
        def completed_extraction(attachment):
            return BillExtraction.objects.create(
                attachment=attachment,
                status=BillExtraction.Status.COMPLETED,
                raw_text="Rapido ride invoice total 245",
                extracted_vendor="Rapido",
                extracted_date=date(2026, 6, 24),
                extracted_amount=Decimal("245.00"),
            )

        mocked_ocr.side_effect = completed_extraction

        result = enqueue_smart_upload(self.upload.id)

        self.assertEqual(result.get()["status"], SmartReimbursementUpload.Status.READY_TO_CONFIRM)
        self.upload.refresh_from_db()
        self.assertEqual(self.upload.status, SmartReimbursementUpload.Status.READY_TO_CONFIRM)
        self.assertEqual(self.upload.processed_files, 1)

    @patch("reimbursements.email_services.EmailMessage")
    def test_confirm_and_send_materializes_reports_then_sends_audited_email(self, mocked_email_class):
        self._prepare_confirmable_upload()
        mocked_email_class.return_value.send.return_value = 1

        log = confirm_and_send_upload(self.upload.id, self.user.id)

        self.assertEqual(log.status, EmailDispatchLog.Status.SENT)
        self.assertEqual(log.smart_reimbursement_upload_id, self.upload.id)
        self.assertEqual(log.triggered_by, self.user)
        self.assertEqual(log.to_email, "finance@vbsai.com")
        self.assertEqual(log.cc_emails, ["manager@vbsai.com", "hr@vbsai.com"])
        self.assertEqual(log.attached_reports.count(), 2)
        self.assertEqual(
            set(log.attached_reports.values_list("report_type", flat=True)),
            {GeneratedReport.ReportType.QUICK_CLAIM_EXCEL, GeneratedReport.ReportType.QUICK_CLAIM_PDF},
        )
        self.assertEqual(log.subject, "Reimbursement Claim Report - Jithin Raj - June 2026")
        self.assertIn("Final Confirmed Amount: INR 245.00", log.body)
        self.assertIn("generated only after the employee reviewed and verified", log.body)
        self.assertIn("ReimburIQ Automation", log.body)
        self.upload.refresh_from_db()
        self.assertEqual(self.upload.status, SmartReimbursementUpload.Status.SENT)
        self.assertEqual(self.upload.reimbursement_profile, self.profile)
        self.assertIsNotNone(self.upload.sent_at)
        self.assertEqual(self.upload.claim.source, "QUICK_BULK_UPLOAD")
        self.assertEqual(self.upload.claim.expense_items.count(), 1)
        self.attachment.refresh_from_db()
        self.assertEqual(self.attachment.expense_item, self.upload.claim.expense_items.get())
        self.assertIsNotNone(self.upload.excel_report_id)
        self.assertIsNotNone(self.upload.pdf_report_id)
        mocked_email = mocked_email_class.return_value
        self.assertEqual(mocked_email.attach_file.call_count, 2)
        attached_paths = [call.args[0] for call in mocked_email.attach_file.call_args_list]
        self.assertTrue(any(path.endswith(".xlsx") for path in attached_paths))
        self.assertTrue(any(path.endswith(".pdf") for path in attached_paths))
        mocked_email.send.assert_called_once_with(fail_silently=False)

    @patch("reimbursements.email_services.EmailMessage")
    def test_quick_claim_reports_are_formatted_and_do_not_expose_absolute_paths(self, mocked_email_class):
        self._prepare_confirmable_upload()
        mocked_email_class.return_value.send.return_value = 1

        confirm_and_send_upload(self.upload.id, self.user.id)
        self.upload.refresh_from_db()

        self.assertEqual(self.upload.excel_report.report_type, GeneratedReport.ReportType.QUICK_CLAIM_EXCEL)
        self.assertEqual(self.upload.pdf_report.report_type, GeneratedReport.ReportType.QUICK_CLAIM_PDF)
        workbook = load_workbook(self.upload.excel_report.file.path)
        sheet = workbook["Quick Claim Summary"]
        cell_values = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None]
        report_text = "\n".join(cell_values)
        self.assertIn("Quick Claim Reimbursement Report", report_text)
        self.assertIn("NL Technologies Pvt. Ltd.", report_text)
        self.assertIn("Jithin Raj", report_text)
        self.assertIn("June 2026", report_text)
        self.assertIn("Uploaded / Processed / Failed", report_text)
        self.assertIn("1 / 1 / 0", report_text)
        self.assertIn("Manually Edited Rows", report_text)
        self.assertIn("Final Confirmed Amount", report_text)
        self.assertIn("rapido.png", report_text)
        self.assertIn("Final Confirmed Total", report_text)
        self.assertNotIn(str(Path(TEST_MEDIA_ROOT).resolve()), report_text)
        self.assertEqual(sheet.freeze_panes, "A15")
        headers = [sheet.cell(row=14, column=column).value for column in range(1, 12)]
        self.assertEqual(
            headers,
            [
                "S.No",
                "Bill File Name",
                "Bill Date",
                "Vendor",
                "Purpose / Details of Claim",
                "Category",
                "Amount",
                "Classification Confidence",
                "Source",
                "Review Status",
                "Remarks",
            ],
        )
        self.assertEqual(sheet.cell(row=15, column=7).number_format, "#,##0.00")
        self.assertGreater(self.upload.pdf_report.file.size, 500)
        with self.upload.pdf_report.file.open("rb") as pdf_file:
            pdf_bytes = pdf_file.read()
            self.assertNotIn(str(Path(TEST_MEDIA_ROOT).resolve()).encode(), pdf_bytes)
        pdf_document = fitz.open(self.upload.pdf_report.file.path)
        try:
            pdf_text = "\n".join(page.get_text() for page in pdf_document)
        finally:
            pdf_document.close()
        self.assertIn("Jithin Raj", pdf_text)
        self.assertIn("June 2026", pdf_text)
        self.assertIn("Final Confirmed Amount", pdf_text)
        self.assertIn("Travel from office to client meeting", pdf_text)

    @patch("reimbursements.email_services.EmailMessage")
    def test_duplicate_confirm_does_not_duplicate_items_reports_or_email(self, mocked_email_class):
        self._prepare_confirmable_upload()
        mocked_email_class.return_value.send.return_value = 1

        confirm_and_send_upload(self.upload.id, self.user.id)
        with self.assertRaises(ValidationError):
            confirm_and_send_upload(self.upload.id, self.user.id)

        self.upload.refresh_from_db()
        self.assertEqual(self.upload.claim.expense_items.count(), 1)
        self.assertEqual(GeneratedReport.objects.filter(claim=self.upload.claim).count(), 2)
        self.assertEqual(EmailDispatchLog.objects.filter(smart_reimbursement_upload=self.upload).count(), 1)

    @patch("reimbursements.email_services.EmailMessage")
    def test_confirm_and_send_succeeds_for_vbsai_recipient_and_cc(self, mocked_email_class):
        self._prepare_confirmable_upload()
        self.profile.finance_head_email = "rahul.ab@vbsai.com"
        self.profile.cc_emails = ["jithin.raj@vbsai.com"]
        self.profile.save(update_fields=["finance_head_email", "cc_emails", "updated_at"])
        mocked_email_class.return_value.send.return_value = 1

        log = confirm_and_send_upload(self.upload.id, self.user.id)

        self.assertEqual(log.status, EmailDispatchLog.Status.SENT)
        self.assertEqual(log.to_email, "rahul.ab@vbsai.com")
        self.assertEqual(log.cc_emails, ["jithin.raj@vbsai.com"])
        mocked_email_class.return_value.send.assert_called_once_with(fail_silently=False)


class QuickClaimRecipientPolicyTests(TestCase):
    @override_settings(ALLOWED_RECIPIENT_DOMAINS=[])
    def test_recipient_policy_requires_configured_allowlist(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=[])
        self.assertFalse(recipient_is_allowed("finance@vbsai.com"))

    def test_recipient_policy_accepts_exact_and_subdomain_matches(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])
        self.assertTrue(recipient_is_allowed("finance@vbsai.com"))
        self.assertTrue(recipient_is_allowed("finance@internal.vbsai.com"))
        self.assertFalse(recipient_is_allowed("finance@vbsai.com.attacker.example"))

    def test_recipient_policy_accepts_vbsai_finance_and_cto_recipients(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])

        self.assertTrue(recipient_is_allowed("rahul.ab@vbsai.com"))
        self.assertTrue(recipient_is_allowed("jithin.raj@vbsai.com"))

    def test_recipient_policy_rejects_gmail_when_only_vbsai_is_allowlisted(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])

        self.assertFalse(recipient_is_allowed("personal@gmail.com"))
        with self.assertRaisesMessage(
            ValidationError,
            "Recipient domain is not allowlisted. Add vbsai.com in System Settings.",
        ):
            validate_recipients("rahul.ab@vbsai.com", ["personal@gmail.com"])

    def test_full_email_accidentally_stored_in_allowlist_normalizes_to_domain(self):
        setting = SystemSetting.objects.create(
            quick_claim_allowed_recipient_domains=[
                " Rahul.Ab@VbsAI.com ",
                "vbsai.com",
                "",
            ]
        )

        setting.refresh_from_db()
        self.assertEqual(setting.quick_claim_allowed_recipient_domains, ["vbsai.com"])
        self.assertTrue(recipient_is_allowed("jithin.raj@vbsai.com"))

    @override_settings(ALLOWED_RECIPIENT_DOMAINS=["gmail.com"])
    def test_db_allowlist_has_priority_over_env_fallback(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=["vbsai.com"])

        self.assertTrue(recipient_is_allowed("rahul.ab@vbsai.com"))
        self.assertFalse(recipient_is_allowed("person@gmail.com"))

    @override_settings(ALLOWED_RECIPIENT_DOMAINS=["gmail.com"])
    def test_empty_db_allowlist_uses_env_fallback(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=[])

        self.assertTrue(recipient_is_allowed("person@gmail.com"))
        self.assertFalse(recipient_is_allowed("rahul.ab@vbsai.com"))

    @override_settings(ALLOWED_RECIPIENT_DOMAINS=[])
    def test_empty_allowlist_gives_clear_error_before_smtp(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=[])

        with self.assertRaisesMessage(
            ValidationError,
            "No Quick Claim recipient domains are configured. Add vbsai.com in System Settings.",
        ):
            validate_recipients("rahul.ab@vbsai.com", ["jithin.raj@vbsai.com"])


class EmployeeReimbursementProfileApiTests(APITestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="profile-user",
            email="profile@example.com",
            password="secret-pass",
        )
        self.client.force_authenticate(self.user)

    @override_settings(ALLOWED_RECIPIENT_DOMAINS=["vbsai.com"])
    def test_get_missing_profile_returns_incomplete_payload(self):
        response = self.client.get(reverse("my-reimbursement-profile"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data["is_complete"])
        self.assertEqual(response.data["email"], "profile@example.com")

    @override_settings(ALLOWED_RECIPIENT_DOMAINS=["vbsai.com"])
    def test_profile_save_validates_required_fields_month_and_year(self):
        response = self.client.put(
            reverse("my-reimbursement-profile"),
            {
                "employee_name": "",
                "department": "",
                "default_claim_month": 13,
                "default_claim_year": 2023,
                "finance_head_email": "rahul.ab@vbsai.com",
                "cc_emails": [],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("employee_name", response.data)
        self.assertIn("department", response.data)
        self.assertIn("default_claim_month", response.data)
        self.assertIn("default_claim_year", response.data)

    @override_settings(ALLOWED_RECIPIENT_DOMAINS=["vbsai.com"])
    def test_profile_save_validates_finance_and_cc_allowed_domains(self):
        finance_response = self.client.put(
            reverse("my-reimbursement-profile"),
            {
                "employee_name": "Jithin Raj",
                "department": "Technology",
                "default_claim_month": 6,
                "default_claim_year": 2026,
                "finance_head_email": "finance@external.example",
                "cc_emails": [],
            },
            format="json",
        )
        cc_response = self.client.put(
            reverse("my-reimbursement-profile"),
            {
                "employee_name": "Jithin Raj",
                "department": "Technology",
                "default_claim_month": 6,
                "default_claim_year": 2026,
                "finance_head_email": "rahul.ab@vbsai.com",
                "cc_emails": ["manager@external.example"],
            },
            format="json",
        )

        self.assertEqual(finance_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(cc_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Recipient domain is not allowlisted", str(finance_response.data))
        self.assertIn("Recipient domain is not allowlisted", str(cc_response.data))

    @override_settings(ALLOWED_RECIPIENT_DOMAINS=["gmail.com"])
    def test_profile_save_uses_env_fallback_when_db_allowlist_empty(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=[])

        response = self.client.put(
            reverse("my-reimbursement-profile"),
            {
                "employee_name": "Jithin Raj",
                "department": "Technology",
                "default_claim_month": 6,
                "default_claim_year": 2026,
                "finance_head_email": "employee@gmail.com",
                "cc_emails": [],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data["is_complete"])

    @override_settings(ALLOWED_RECIPIENT_DOMAINS=[])
    def test_profile_save_blocks_when_db_and_env_allowlist_empty(self):
        SystemSetting.objects.create(quick_claim_allowed_recipient_domains=[])

        response = self.client.put(
            reverse("my-reimbursement-profile"),
            {
                "employee_name": "Jithin Raj",
                "department": "Technology",
                "default_claim_month": 6,
                "default_claim_year": 2026,
                "finance_head_email": "rahul.ab@vbsai.com",
                "cc_emails": [],
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("No Quick Claim recipient domains are configured", str(response.data))
