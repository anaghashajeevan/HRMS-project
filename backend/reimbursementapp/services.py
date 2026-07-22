import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Sum
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import load_workbook
from xlrd import XL_CELL_BLANK, XL_CELL_BOOLEAN, XL_CELL_DATE, XL_CELL_EMPTY, XL_CELL_NUMBER
import xlrd

from HRMSapp.models import Employee

from .models import ExpenseItem, MonthlyReimbursementBatch, ReimbursementClaim, UploadedReimbursementForm

MONTH_PATTERN = re.compile(
    r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\b",
    re.IGNORECASE,
)
YEAR_PATTERN = re.compile(r"\b(20\d{2}|19\d{2})\b")

CATEGORY_COLUMN_CONFIG = [
    ("local_transport_taxi", "Local Transport / Taxi", ["local", "transport", "taxi"]),
    ("meal", "Meal", ["meal"]),
    ("telephone", "Telephone", ["telephone"]),
    ("travelling_hotel", "Travelling / Hotel", ["travelling", "travel", "hotel"]),
    ("office_expenses", "Office Expenses", ["office", "expenses"]),
    ("others_misc", "Others / Misc", ["others", "misc"]),
]

CATEGORY_IMPORT_MAP = {
    "Office Expenses": ExpenseItem.Category.OFFICE_SUPPLIES,
    "Others / Misc": ExpenseItem.Category.OTHER,
    "Local Transport / Taxi": ExpenseItem.Category.TRAVEL,
    "Meal": ExpenseItem.Category.FOOD,
    "Telephone": ExpenseItem.Category.OTHER,
    "Travelling / Hotel": ExpenseItem.Category.TRAVEL,
}


def _format_datetime_value(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def _normalize_text(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower()).strip()


def _to_primitive(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return _format_datetime_value(value)
    if isinstance(value, Decimal):
        number = float(value)
        return int(number) if number.is_integer() else number
    if isinstance(value, float):
        return int(value) if value.is_integer() else value
    if isinstance(value, (int, bool)):
        return value
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _normalize_xlrd_cell(cell, datemode):
    if cell.ctype == XL_CELL_DATE:
        try:
            return _to_primitive(xlrd.xldate.xldate_as_datetime(cell.value, datemode))
        except Exception:
            return str(cell.value)
    if cell.ctype == XL_CELL_NUMBER:
        return _to_primitive(float(cell.value))
    if cell.ctype == XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in (XL_CELL_EMPTY, XL_CELL_BLANK):
        return ""
    return _to_primitive(cell.value)


def _read_xlsx_rows(file_path: str):
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
    except Exception as exc:
        raise ValueError(f"Failed to parse reimbursement form: {exc}") from exc

    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([_to_primitive(cell) for cell in row])

    return {
        "file_type": ".xlsx",
        "sheet_name": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "rows": rows,
    }


def _read_xls_rows(file_path: str):
    try:
        workbook = xlrd.open_workbook(file_path, formatting_info=False)
    except Exception as exc:
        raise ValueError(f"Failed to parse reimbursement form: {exc}") from exc

    sheet = workbook.sheet_by_index(0)
    rows = []
    for row_index in range(sheet.nrows):
        rows.append([_normalize_xlrd_cell(cell, workbook.datemode) for cell in sheet.row(row_index)])

    return {
        "file_type": ".xls",
        "sheet_name": sheet.name,
        "max_row": sheet.nrows,
        "max_column": sheet.ncols,
        "rows": rows,
    }


def _read_workbook_rows(file_path: str):
    suffix = Path(file_path).suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx_rows(file_path)
    if suffix == ".xls":
        return _read_xls_rows(file_path)
    raise ValueError(f"Unsupported reimbursement form format: {suffix or 'unknown'}.")


def _clean_rows(rows):
    cleaned_rows = []
    for row in rows:
        normalized_row = [_to_primitive(cell) for cell in row]
        if any(cell != "" for cell in normalized_row):
            cleaned_rows.append(normalized_row)
    return cleaned_rows


def _find_label_value(rows, label):
    normalized_label = _normalize_text(label)
    for row in rows:
        for index, cell in enumerate(row):
            if _normalize_text(cell) == normalized_label:
                for right_cell in row[index + 1 :]:
                    right_text = str(right_cell).strip()
                    if right_text and right_text not in {":", "-", "--"}:
                        return str(right_cell).strip(), row
    return "", []


def _extract_month_year(text):
    month_match = MONTH_PATTERN.search(text or "")
    year_match = YEAR_PATTERN.search(text or "")
    month = month_match.group(1).title() if month_match else ""
    year = year_match.group(1) if year_match else ""
    return month, year


def _extract_employee_name(text):
    cleaned = MONTH_PATTERN.sub("", text or "")
    cleaned = YEAR_PATTERN.sub("", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .:-")
    return cleaned


def _parse_numeric(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return None
        try:
            number = float(cleaned)
            return int(number) if number.is_integer() else number
        except ValueError:
            return None
    return None


def _parse_serial_number(value):
    number = _parse_numeric(value)
    return int(number) if isinstance(number, (int, float)) else None


def _find_expense_header_row(rows):
    for row_index, row in enumerate(rows):
        normalized_cells = [_normalize_text(cell) for cell in row]
        joined = " ".join(normalized_cells)
        if "s n" in joined and "receipt inv date" in joined and "purpose details of claim" in joined:
            return row_index, normalized_cells
    raise ValueError("Could not locate the expense table header in the uploaded form.")


def _find_column_index(normalized_cells, required_terms):
    for index, cell in enumerate(normalized_cells):
        if all(term in cell for term in required_terms):
            return index
    return None


def _build_column_map(normalized_header_cells):
    column_map = {
        "serial_no": _find_column_index(normalized_header_cells, ["s", "n"]),
        "receipt_date": _find_column_index(normalized_header_cells, ["receipt", "date"]),
        "purpose": _find_column_index(normalized_header_cells, ["purpose", "claim"]),
        "grand_total": _find_column_index(normalized_header_cells, ["grand", "total", "inr"]),
    }

    for key, _, terms in CATEGORY_COLUMN_CONFIG:
        column_map[key] = _find_column_index(normalized_header_cells, terms)

    if column_map["serial_no"] is None or column_map["receipt_date"] is None or column_map["purpose"] is None:
        raise ValueError("The uploaded form is missing one or more required expense table columns.")

    return column_map


def _row_value(row, index):
    if index is None or index >= len(row):
        return ""
    return row[index]


def _determine_category_and_amount(row, column_map):
    raw_category_values = {}
    selected_category = "Other"
    selected_amount = None

    for key, label, _ in CATEGORY_COLUMN_CONFIG:
        value = _row_value(row, column_map.get(key))
        numeric_value = _parse_numeric(value)
        raw_category_values[key] = numeric_value if numeric_value is not None else (value if value != "" else "")
        if selected_amount is None and numeric_value is not None:
            selected_category = label
            selected_amount = numeric_value

    grand_total_value = _parse_numeric(_row_value(row, column_map.get("grand_total")))
    amount = grand_total_value if grand_total_value is not None else selected_amount
    return selected_category, amount, raw_category_values


def parse_company_reimbursement_form(file_path: str) -> dict:
    workbook_data = _read_workbook_rows(file_path)
    rows = _clean_rows(workbook_data["rows"])
    warnings = []

    company_value, _ = _find_label_value(rows, "COMPANY")
    name_value, name_row = _find_label_value(rows, "Name")
    department_value, _ = _find_label_value(rows, "Department")

    month, year = _extract_month_year(name_value)
    if not month or not year:
        joined_name_row = " ".join(str(cell) for cell in name_row if str(cell).strip())
        row_month, row_year = _extract_month_year(joined_name_row)
        month = month or row_month
        year = year or row_year

    employee_name = _extract_employee_name(name_value)

    header_row_index, normalized_header_cells = _find_expense_header_row(rows)
    column_map = _build_column_map(normalized_header_cells)

    expenses = []
    grand_total = None

    for row in rows[header_row_index + 1 :]:
        joined_row = " ".join(_normalize_text(cell) for cell in row if str(cell).strip())
        if any(stop_text in joined_row for stop_text in ["claimer s signature", "signature"]):
            break

        if "total" in joined_row:
            total_value = _parse_numeric(_row_value(row, column_map.get("grand_total")))
            grand_total = total_value if total_value is not None else grand_total
            break

        serial_no = _parse_serial_number(_row_value(row, column_map["serial_no"]))
        if serial_no is None:
            continue

        receipt_date = _row_value(row, column_map["receipt_date"])
        purpose = str(_row_value(row, column_map["purpose"])).strip()
        category, amount, raw_category_values = _determine_category_and_amount(row, column_map)

        if amount is None:
            warnings.append(f"Skipped row {serial_no} because no amount column had a usable value.")
            continue

        expenses.append(
            {
                "serial_no": serial_no,
                "receipt_date": receipt_date,
                "purpose": purpose,
                "category": category,
                "amount": amount,
                "raw_category_values": raw_category_values,
            }
        )

    if grand_total is None:
        grand_total = sum(_parse_numeric(expense["amount"]) or 0 for expense in expenses)
        warnings.append("Grand total row not found. Calculated grand total from parsed expense amounts.")

    if not company_value:
        warnings.append("Company name could not be identified from the uploaded form.")
    if not employee_name:
        warnings.append("Employee name could not be identified from the uploaded form.")
    if not department_value:
        warnings.append("Department could not be identified from the uploaded form.")
    if not month or not year:
        warnings.append("Month and year could not be confidently identified from the uploaded form.")

    return {
        "parser": "company_reimbursement_v1",
        "file_type": workbook_data["file_type"],
        "sheet_name": workbook_data["sheet_name"],
        "company": company_value,
        "employee_name": employee_name,
        "department": department_value,
        "month": month,
        "year": year,
        "expenses": expenses,
        "grand_total": grand_total,
        "row_count": len(expenses),
        "warnings": warnings,
    }


def parse_reimbursement_excel(file_path: str) -> dict:
    workbook_data = _read_workbook_rows(file_path)
    cleaned_rows = _clean_rows(workbook_data["rows"])
    preview_rows = cleaned_rows[:20]

    return {
        "file_type": workbook_data["file_type"],
        "sheet_name": workbook_data["sheet_name"],
        "max_row": workbook_data["max_row"],
        "max_column": workbook_data["max_column"],
        "preview_rows": preview_rows,
    }


def _generate_employee_code(full_name: str) -> str:
    code = re.sub(r"[^A-Za-z0-9]+", "_", full_name.upper()).strip("_")
    return code or "EMPLOYEE"


def _get_or_create_employee(parsed_data: dict) -> Employee:
    """
    Find or create an employee from parsed reimbursement data.
    Adapted for HRMS Employee model which uses first_name + last_name
    instead of a single full_name field.
    """
    full_name = str(parsed_data.get("employee_name", "")).strip()
    if not full_name:
        raise ValueError("Parsed reimbursement form does not contain a usable employee name.")

    department = str(parsed_data.get("department", "")).strip()
    
    # Split full_name into first_name and last_name
    name_parts = full_name.split(None, 1)  # Split on first space
    first_name = name_parts[0] if name_parts else full_name
    last_name = name_parts[1] if len(name_parts) > 1 else ''

    # Try to find existing employee by first_name + last_name
    existing_employee = Employee.objects.filter(
        first_name__iexact=first_name,
        last_name__iexact=last_name,
        is_deleted=False,
    ).first()
    
    # Also try by official email if provided
    if not existing_employee:
        email = str(parsed_data.get("email", "")).strip()
        if email:
            existing_employee = Employee.objects.filter(
                official_email__iexact=email,
                is_deleted=False,
            ).first()

    if existing_employee:
        return existing_employee

    # Generate employee_id
    base_code = _generate_employee_code(full_name)
    candidate_code = base_code
    suffix = 1
    while Employee.objects.filter(employee_id=candidate_code).exists():
        suffix += 1
        candidate_code = f"{base_code}_{suffix}"

    # Create new employee
    # Note: HRMS Employee requires date_of_birth, date_of_joining, phone_number
    # Use placeholder values for auto-created employees
    from django.utils import timezone
    
    return Employee.objects.create(
        employee_id=candidate_code,
        first_name=first_name,
        last_name=last_name,
        official_email=f"{candidate_code.lower()}@placeholder.local",
        phone_number="0000000000",
        date_of_birth=timezone.now().date().replace(year=2000),
        date_of_joining=timezone.now().date(),
        status='ACTIVE',
    )


def _month_name_to_number(month_name: str) -> int:
    if not month_name:
        raise ValueError("Parsed reimbursement form does not contain a month.")

    for index in range(1, 13):
        if datetime(2000, index, 1).strftime("%B").lower() == month_name.lower():
            return index
    raise ValueError(f"Could not convert month '{month_name}' to a month number.")


def _get_or_create_batch(parsed_data: dict) -> MonthlyReimbursementBatch:
    month_name = str(parsed_data.get("month", "")).strip()
    year_value = str(parsed_data.get("year", "")).strip()
    if not year_value.isdigit():
        raise ValueError("Parsed reimbursement form does not contain a valid year.")

    month_number = _month_name_to_number(month_name)
    year_number = int(year_value)
    title = f"{month_name.title()} {year_number} Reimbursement Batch"

    batch, created = MonthlyReimbursementBatch.objects.get_or_create(
        month=month_number,
        year=year_number,
        defaults={"title": title},
    )
    if not created and not batch.title:
        batch.title = title
        batch.save(update_fields=["title", "updated_at"])
    return batch


def _parse_expense_date(value) -> date | None:
    if not value:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()

    text = str(value).strip()
    parsed = parse_date(text)
    if parsed:
        return parsed

    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _expense_category_from_parsed(label: str) -> str:
    return CATEGORY_IMPORT_MAP.get(label, ExpenseItem.Category.OTHER)


def _to_decimal_amount(value) -> Decimal:
    numeric = value
    if isinstance(value, str):
        numeric = value.replace(",", "").strip()
    return Decimal(str(numeric or 0)).quantize(Decimal("0.01"))


def _recalculate_batch_totals(batch: MonthlyReimbursementBatch) -> None:
    batch.total_employees = batch.claims.count()
    totals = batch.claims.aggregate(
        total_claimed=Sum("total_claimed_amount"),
        total_approved=Sum("total_approved_amount"),
    )
    batch.total_claimed_amount = totals["total_claimed"] or Decimal("0.00")
    batch.total_approved_amount = totals["total_approved"] or Decimal("0.00")
    batch.save(
        update_fields=[
            "total_employees",
            "total_claimed_amount",
            "total_approved_amount",
            "updated_at",
        ]
    )


def recalculate_claim_and_batch_totals(claim: ReimbursementClaim) -> ReimbursementClaim:
    approved_totals = claim.expense_items.filter(status=ExpenseItem.Status.APPROVED).aggregate(
        total_approved=Sum("approved_amount")
    )
    claim.total_approved_amount = approved_totals["total_approved"] or Decimal("0.00")
    claim.save(update_fields=["total_approved_amount", "updated_at"])
    _recalculate_batch_totals(claim.batch)
    return claim


@transaction.atomic
def approve_expense_item(
    expense_item: ExpenseItem,
    approved_amount: Decimal | None = None,
    review_notes: str = "",
) -> ExpenseItem:
    expense_item.status = ExpenseItem.Status.APPROVED
    expense_item.approved_amount = approved_amount if approved_amount is not None else expense_item.claimed_amount
    if review_notes.strip():
        expense_item.review_notes = review_notes.strip()
    expense_item.save(update_fields=["status", "approved_amount", "review_notes", "updated_at"])
    recalculate_claim_and_batch_totals(expense_item.claim)
    return expense_item


@transaction.atomic
def reject_expense_item(expense_item: ExpenseItem, review_notes: str) -> ExpenseItem:
    expense_item.status = ExpenseItem.Status.REJECTED
    expense_item.approved_amount = Decimal("0.00")
    expense_item.review_notes = review_notes.strip()
    expense_item.save(update_fields=["status", "approved_amount", "review_notes", "updated_at"])
    recalculate_claim_and_batch_totals(expense_item.claim)
    return expense_item


@transaction.atomic
def approve_matched_claim_expenses(claim: ReimbursementClaim) -> dict:
    approved_count = 0
    skipped_count = 0

    for expense_item in claim.expense_items.select_related("validation").all():
        try:
            validation = expense_item.validation
        except ObjectDoesNotExist:
            validation = None
        if validation and validation.status == "MATCHED":
            expense_item.status = ExpenseItem.Status.APPROVED
            expense_item.approved_amount = expense_item.claimed_amount
            expense_item.review_notes = "Auto-approved matched bill."
            expense_item.save(update_fields=["status", "approved_amount", "review_notes", "updated_at"])
            approved_count += 1
        else:
            skipped_count += 1

    claim = recalculate_claim_and_batch_totals(claim)

    return {
        "claim_id": claim.id,
        "approved_count": approved_count,
        "skipped_count": skipped_count,
        "total_approved_amount": f"{claim.total_approved_amount:.2f}",
    }


@transaction.atomic
def import_parsed_reimbursement_form(uploaded_form: UploadedReimbursementForm) -> ReimbursementClaim:
    if uploaded_form.imported_claim_id:
        raise ValueError("This form has already been imported into a claim.")

    parsed_data = uploaded_form.parsed_data or {}
    if not parsed_data:
        raise ValueError("This uploaded form does not have parsed reimbursement data yet.")
    if parsed_data.get("parser") != "company_reimbursement_v1":
        raise ValueError("Only company reimbursement parser output can be imported into a claim.")

    expenses = parsed_data.get("expenses") or []
    if not expenses:
        raise ValueError("No parsed expense rows were found to import.")

    employee = _get_or_create_employee(parsed_data)
    batch = _get_or_create_batch(parsed_data)

    claim, created = ReimbursementClaim.objects.get_or_create(
        batch=batch,
        employee=employee,
        defaults={
            "status": ReimbursementClaim.Status.SUBMITTED,
            "remarks": f"Imported from uploaded reimbursement form: {uploaded_form.original_filename}",
        },
    )

    if not created and claim.expense_items.exists():
        raise ValueError("This form has already been imported into a claim.")

    total_claimed_amount = Decimal("0.00")
    if not created:
        claim.status = ReimbursementClaim.Status.SUBMITTED
        claim.remarks = f"Imported from uploaded reimbursement form: {uploaded_form.original_filename}"
        claim.save(update_fields=["status", "remarks", "updated_at"])

    for parsed_expense in expenses:
        amount = _to_decimal_amount(parsed_expense.get("amount", 0))
        total_claimed_amount += amount
        ExpenseItem.objects.create(
            claim=claim,
            expense_date=_parse_expense_date(parsed_expense.get("receipt_date")),
            category=_expense_category_from_parsed(str(parsed_expense.get("category", "")).strip()),
            vendor_name="",
            description=str(parsed_expense.get("purpose", "")).strip(),
            claimed_amount=amount,
            approved_amount=Decimal("0.00"),
            status=ExpenseItem.Status.PENDING_REVIEW,
        )

    if total_claimed_amount == Decimal("0.00"):
        total_claimed_amount = _to_decimal_amount(parsed_data.get("grand_total", 0))

    claim.total_claimed_amount = total_claimed_amount
    claim.total_approved_amount = Decimal("0.00")
    claim.status = ReimbursementClaim.Status.SUBMITTED
    claim.remarks = f"Imported from uploaded reimbursement form: {uploaded_form.original_filename}"
    claim.save(
        update_fields=[
            "total_claimed_amount",
            "total_approved_amount",
            "status",
            "remarks",
            "updated_at",
        ]
    )

    _recalculate_batch_totals(batch)

    uploaded_form.employee = employee
    uploaded_form.batch = batch
    uploaded_form.imported_claim = claim
    uploaded_form.imported_at = timezone.now()
    uploaded_form.status = UploadedReimbursementForm.Status.IMPORTED
    uploaded_form.save(
        update_fields=[
            "employee",
            "batch",
            "imported_claim",
            "imported_at",
            "status",
        ]
    )

    return claim

