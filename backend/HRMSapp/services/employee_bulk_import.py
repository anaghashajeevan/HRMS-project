# """
# Bulk import HRMS employees from CSV/XLSX.
# """

# import csv
# import re
# from datetime import datetime
# from io import TextIOWrapper

# from openpyxl import load_workbook


# REQUIRED_COLUMNS = [
#     'Employee Code',
#     'First Name',
#     'Last Name',
#     'Email',
#     'Phone',
#     'Date of Joining',
#     'Date of Birth',
# ]

# OPTIONAL_COLUMNS = [
#     'Department',
#     'Gender',
#     'Personal Email',
#     'Position',
# ]


# class EmployeeBulkImportError(Exception):
#     pass


# def _normalize_employee_code(value: str) -> str:
#     """Match the same normalization used in attendance."""
#     if not value:
#         return ""
#     return str(value).strip().replace("-", "").replace(" ", "").upper()


# def _parse_date(value, field_name: str, row_num: int):
#     """Parse date from multiple formats."""
#     if not value:
#         raise ValueError(f"Row {row_num}: {field_name} is required")

#     value_str = str(value).strip()
#     if not value_str:
#         raise ValueError(f"Row {row_num}: {field_name} is required")

#     # Try common date formats
#     formats = ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y']
#     for fmt in formats:
#         try:
#             return datetime.strptime(value_str, fmt).date()
#         except ValueError:
#             continue

#     # Openpyxl might return date object already
#     if hasattr(value, 'date'):
#         return value.date()

#     raise ValueError(
#         f"Row {row_num}: {field_name} '{value_str}' is invalid. "
#         f"Use format YYYY-MM-DD (e.g., 2024-01-15)"
#     )


# def _parse_gender(value):
#     """Normalize gender value."""
#     if not value:
#         return None
#     v = str(value).strip().upper()
#     if v in ['M', 'MALE']:
#         return 'MALE'
#     if v in ['F', 'FEMALE']:
#         return 'FEMALE'
#     if v in ['O', 'OTHER', 'OTHERS']:
#         return 'OTHER'
#     return None


# def _validate_email(email: str) -> bool:
#     """Simple email validation."""
#     if not email:
#         return False
#     pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
#     return bool(re.match(pattern, email.strip()))


# def _read_csv(uploaded_file):
#     """Read CSV file and return list of dict rows."""
#     uploaded_file.seek(0)
#     wrapper = TextIOWrapper(uploaded_file.file, encoding='utf-8-sig', newline='')
#     reader = csv.DictReader(wrapper)
#     return list(reader)


# def _read_xlsx(uploaded_file):
#     """Read XLSX file and return list of dict rows."""
#     uploaded_file.seek(0)
#     workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
#     worksheet = workbook.active
#     rows = worksheet.iter_rows(values_only=True)

#     # First row = headers
#     headers_row = next(rows, None)
#     if not headers_row:
#         return []

#     headers = [str(h).strip() if h else '' for h in headers_row]

#     result = []
#     for row_values in rows:
#         # Skip completely empty rows
#         if not any(v for v in row_values):
#             continue
#         row_dict = {}
#         for i, header in enumerate(headers):
#             if header:
#                 row_dict[header] = row_values[i] if i < len(row_values) else None
#         result.append(row_dict)
#     return result


# def _read_file(uploaded_file):
#     """Auto-detect CSV vs XLSX and read."""
#     filename = uploaded_file.name.lower()
#     if filename.endswith('.csv'):
#         return _read_csv(uploaded_file)
#     elif filename.endswith('.xlsx') or filename.endswith('.xls'):
#         return _read_xlsx(uploaded_file)
#     raise EmployeeBulkImportError(
#         "Unsupported file type. Please upload a CSV or XLSX file."
#     )


# def _validate_headers(rows):
#     """Ensure all required columns are present."""
#     if not rows:
#         raise EmployeeBulkImportError("File is empty or has no data rows.")

#     columns = set(rows[0].keys())
#     missing = [col for col in REQUIRED_COLUMNS if col not in columns]
#     if missing:
#         raise EmployeeBulkImportError(
#             f"Missing required columns: {', '.join(missing)}"
#         )


# def import_employees_from_file(uploaded_file, skip_existing: bool = True) -> dict:
#     """
#     Bulk import HRMS employees from CSV/XLSX.

#     Args:
#         uploaded_file: Django UploadedFile
#         skip_existing: If True, skip existing employees. If False, update them.

#     Returns:
#         dict with: created, updated, skipped, errors, total_rows
#     """
#     from HRMSapp.models import Employee, CompanyStructure, JobPosition

#     rows = _read_file(uploaded_file)
#     _validate_headers(rows)

#     result = {
#         'total_rows': len(rows),
#         'created': 0,
#         'updated': 0,
#         'skipped': 0,
#         'errors': [],
#         'created_employees': [],
#     }

#     # Cache departments and positions for faster lookup
#     departments_cache = {
#         c.name.lower(): c
#         for c in CompanyStructure.objects.filter(type='DEPARTMENT', is_active=True)
#     }
#     positions_cache = {
#         p.title.lower(): p
#         for p in JobPosition.objects.filter(is_active=True)
#     }

#     for idx, row in enumerate(rows, start=2):  # start=2 (row 1 = header)
#         try:
#             # Extract required fields
#             employee_code_raw = str(row.get('Employee Code') or '').strip()
#             if not employee_code_raw:
#                 result['errors'].append(f"Row {idx}: Employee Code is required")
#                 continue

#             employee_code = _normalize_employee_code(employee_code_raw)
#             first_name = str(row.get('First Name') or '').strip()
#             last_name = str(row.get('Last Name') or '').strip()
#             email = str(row.get('Email') or '').strip().lower()
#             phone = str(row.get('Phone') or '').strip()

#             # Validate required text fields
#             if not first_name:
#                 result['errors'].append(f"Row {idx}: First Name is required")
#                 continue
#             if not last_name:
#                 result['errors'].append(f"Row {idx}: Last Name is required")
#                 continue
#             if not _validate_email(email):
#                 result['errors'].append(f"Row {idx}: Invalid email '{email}'")
#                 continue
#             if not phone:
#                 result['errors'].append(f"Row {idx}: Phone is required")
#                 continue

#             # Parse dates
#             date_of_joining = _parse_date(
#                 row.get('Date of Joining'), 'Date of Joining', idx
#             )
#             date_of_birth = _parse_date(
#                 row.get('Date of Birth'), 'Date of Birth', idx
#             )

#             # Optional fields
#             gender = _parse_gender(row.get('Gender'))
#             personal_email = str(row.get('Personal Email') or '').strip().lower() or None

#             # Department (optional — resolve by name)
#             department_name = str(row.get('Department') or '').strip()
#             structure_location = None
#             if department_name:
#                 structure_location = departments_cache.get(department_name.lower())
#                 if not structure_location:
#                     result['errors'].append(
#                         f"Row {idx}: Department '{department_name}' not found. "
#                         f"Create it in Settings > Departments first."
#                     )
#                     continue

#             # Position (optional — resolve by title)
#             position_title = str(row.get('Position') or '').strip()
#             position = None
#             if position_title:
#                 position = positions_cache.get(position_title.lower())
#                 if not position:
#                     result['errors'].append(
#                         f"Row {idx}: Position '{position_title}' not found. "
#                         f"Create it in Settings > Job Positions first."
#                     )
#                     continue

#             # Check if employee exists (match by normalized code OR email)
#             existing = Employee.objects.filter(
#                 is_deleted=False
#             ).filter(
#                 # Match by exact ID OR by normalized ID OR by email
#                 # We'll do it manually for normalization
#             )

#             # Try to find match
#             matched_employee = None
#             for emp in Employee.objects.filter(is_deleted=False).only(
#                 'id', 'employee_id', 'official_email'
#             ):
#                 if _normalize_employee_code(emp.employee_id) == employee_code:
#                     matched_employee = emp
#                     break
#                 if emp.official_email.lower() == email:
#                     matched_employee = emp
#                     break

#             if matched_employee:
#                 if skip_existing:
#                     result['skipped'] += 1
#                     continue
#                 # Update existing
#                 matched_employee.first_name = first_name
#                 matched_employee.last_name = last_name
#                 matched_employee.official_email = email
#                 matched_employee.personal_email = personal_email
#                 matched_employee.phone_number = phone
#                 matched_employee.date_of_birth = date_of_birth
#                 matched_employee.date_of_joining = date_of_joining
#                 if gender:
#                     matched_employee.gender = gender
#                 if position:
#                     matched_employee.position = position
#                 if structure_location:
#                     matched_employee.structure_location = structure_location
#                 matched_employee.save()
#                 result['updated'] += 1
#             else:
#                 # Create new employee
#                 new_emp = Employee.objects.create(
#                     employee_id=employee_code_raw,  # Preserve original format (NL001, not normalized)
#                     first_name=first_name,
#                     last_name=last_name,
#                     official_email=email,
#                     personal_email=personal_email,
#                     phone_number=phone,
#                     date_of_birth=date_of_birth,
#                     date_of_joining=date_of_joining,
#                     gender=gender,
#                     position=position,
#                     structure_location=structure_location,
#                     status='ACTIVE',
#                 )
#                 result['created'] += 1
#                 result['created_employees'].append({
#                     'employee_id': new_emp.employee_id,
#                     'full_name': new_emp.full_name,
#                     'email': new_emp.official_email,
#                 })

#         except ValueError as ve:
#             result['errors'].append(str(ve))
#         except Exception as exc:
#             result['errors'].append(f"Row {idx}: {str(exc)}")

#     return result


# def get_sample_csv_content() -> str:
#     """Generate sample CSV template content."""
#     headers = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
#     sample_rows = [
#         [
#             'NL001', 'Lince', 'Kumar', 'lince@company.com', '9876543210',
#             '2024-01-15', '1995-06-20', 'Engineering', 'MALE', '', ''
#         ],
#         [
#             'NL002', 'Nisha', 'Yadav', 'nisha@company.com', '9876543211',
#             '2024-02-10', '1996-03-15', 'Engineering', 'FEMALE', '', ''
#         ],
#         [
#             'NL003', 'Abhijith', 'Rajendran', 'abhijith@company.com', '9876543212',
#             '2024-03-05', '1994-11-08', 'Engineering', 'MALE', '', ''
#         ],
#     ]

#     lines = [','.join(headers)]
#     for row in sample_rows:
#         # Wrap fields containing commas in quotes
#         escaped = [f'"{val}"' if ',' in str(val) else str(val) for val in row]
#         lines.append(','.join(escaped))
#     return '\n'.join(lines)


"""
Bulk import HRMS employees from CSV/XLSX.
Supports two template formats:
1. Simple (5 columns): Employee Code, Employee Name, Department, Email, Status
2. Full (11 columns): with First Name, Last Name, Phone, DOB, DOJ, etc.
"""

import csv
import re
from datetime import datetime, date
from io import TextIOWrapper

from openpyxl import load_workbook


# Simple format (from original attendance project)
SIMPLE_REQUIRED = ['Employee Code', 'Employee Name']
SIMPLE_OPTIONAL = ['Department', 'Email', 'Status']

# Full format (HRMS standard)
FULL_REQUIRED = [
    'Employee Code', 'First Name', 'Last Name', 'Email',
    'Phone', 'Date of Joining', 'Date of Birth',
]
FULL_OPTIONAL = ['Department', 'Gender', 'Personal Email', 'Position']


class EmployeeBulkImportError(Exception):
    pass


def _normalize_employee_code(value: str) -> str:
    if not value:
        return ""
    return str(value).strip().replace("-", "").replace(" ", "").upper()


def _split_full_name(full_name: str) -> tuple[str, str]:
    """
    Split 'Abhijith Rajendran' → ('Abhijith', 'Rajendran')
    Single name like 'Lince' → ('Lince', '') — empty last name (not hyphen)
    """
    name = str(full_name or "").strip()
    if not name:
        return "", ""
    parts = name.split(maxsplit=1)
    if len(parts) == 1:
        return parts[0], ""  # ⬅️ CHANGED: empty string instead of "-"
    return parts[0], parts[1]


def _parse_date(value, field_name: str, row_num: int, default=None):
    """Parse date. Returns default if empty and default provided."""
    if not value or str(value).strip() == "":
        if default is not None:
            return default
        raise ValueError(f"Row {row_num}: {field_name} is required")

    # Openpyxl may return date object
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value

    value_str = str(value).strip()
    formats = ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y']
    for fmt in formats:
        try:
            return datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue

    raise ValueError(
        f"Row {row_num}: {field_name} '{value_str}' is invalid. Use YYYY-MM-DD format."
    )


def _parse_gender(value):
    if not value:
        return None
    v = str(value).strip().upper()
    if v in ['M', 'MALE']:
        return 'MALE'
    if v in ['F', 'FEMALE']:
        return 'FEMALE'
    if v in ['O', 'OTHER', 'OTHERS']:
        return 'OTHER'
    return None


def _parse_status(value):
    """Parse Active/Inactive status → is_active bool + HRMS status string."""
    if not value:
        return True, 'ACTIVE'
    v = str(value).strip().lower()
    if v in ['active', 'true', 'yes', '1']:
        return True, 'ACTIVE'
    if v in ['probation', 'prob']:
        return True, 'PROBATION'
    if v in ['inactive', 'false', 'no', '0', 'terminated', 'left', 'resigned']:
        return False, 'TERMINATED'
    if v in ['suspended']:
        return True, 'SUSPENDED'
    return True, 'ACTIVE'


def _validate_email(email: str) -> bool:
    if not email:
        return False
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email.strip()))


def _read_csv(uploaded_file):
    uploaded_file.seek(0)
    wrapper = TextIOWrapper(uploaded_file.file, encoding='utf-8-sig', newline='')
    reader = csv.DictReader(wrapper)
    return list(reader)


def _read_xlsx(uploaded_file):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers_row = next(rows, None)
    if not headers_row:
        return []
    headers = [str(h).strip() if h else '' for h in headers_row]
    result = []
    for row_values in rows:
        if not any(v for v in row_values):
            continue
        row_dict = {}
        for i, header in enumerate(headers):
            if header:
                row_dict[header] = row_values[i] if i < len(row_values) else None
        result.append(row_dict)
    return result


def _read_file(uploaded_file):
    filename = uploaded_file.name.lower()
    if filename.endswith('.csv'):
        return _read_csv(uploaded_file)
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        return _read_xlsx(uploaded_file)
    raise EmployeeBulkImportError("Unsupported file type. Please upload CSV or XLSX.")


def _detect_format(rows):
    """Detect if it's Simple or Full template based on columns present."""
    if not rows:
        raise EmployeeBulkImportError("File is empty.")

    columns = set(rows[0].keys())

    # Check Full format (has First Name + Last Name separately)
    if 'First Name' in columns and 'Last Name' in columns:
        return 'FULL'

    # Check Simple format (has Employee Name)
    if 'Employee Name' in columns and 'Employee Code' in columns:
        return 'SIMPLE'

    raise EmployeeBulkImportError(
        "Invalid template. Required columns: "
        "either ['Employee Code', 'Employee Name'] (simple) "
        "or ['Employee Code', 'First Name', 'Last Name', 'Email', 'Phone', 'Date of Joining', 'Date of Birth'] (full)"
    )


def import_employees_from_file(uploaded_file, skip_existing: bool = True) -> dict:
    """
    Bulk import HRMS employees from CSV/XLSX.
    Auto-detects Simple vs Full format.
    """
    from HRMSapp.models import Employee, CompanyStructure, JobPosition

    rows = _read_file(uploaded_file)
    template_format = _detect_format(rows)

    result = {
        'template_format': template_format,
        'total_rows': len(rows),
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
        'created_employees': [],
        'warnings': [],
    }

    # Placeholder defaults for missing fields (Simple format)
    PLACEHOLDER_PHONE = '0000000000'
    PLACEHOLDER_DOB = date(1990, 1, 1)
    PLACEHOLDER_DOJ = date.today()

    if template_format == 'SIMPLE':
        result['warnings'].append(
            "Simple template detected. Missing fields (Phone, DOB, DOJ) will use placeholder values. "
            "Please edit each employee's profile in HRMS to add real values."
        )

    # Cache departments and positions
    departments_cache = {
        c.name.lower(): c
        for c in CompanyStructure.objects.filter(type='DEPARTMENT', is_active=True)
    }
    positions_cache = {
        p.title.lower(): p
        for p in JobPosition.objects.filter(is_active=True)
    }

    for idx, row in enumerate(rows, start=2):
        try:
            # ========== COMMON FIELDS ==========
            employee_code_raw = str(row.get('Employee Code') or '').strip()
            if not employee_code_raw:
                result['errors'].append(f"Row {idx}: Employee Code is required")
                continue

            employee_code_normalized = _normalize_employee_code(employee_code_raw)

            # ========== NAME (format-specific) ==========
            if template_format == 'SIMPLE':
                full_name = str(row.get('Employee Name') or '').strip()
                if not full_name:
                    result['errors'].append(f"Row {idx}: Employee Name is required")
                    continue
                first_name, last_name = _split_full_name(full_name)
            else:  # FULL
                first_name = str(row.get('First Name') or '').strip()
                last_name = str(row.get('Last Name') or '').strip()
                if not first_name:
                    result['errors'].append(f"Row {idx}: First Name is required")
                    continue
                if not last_name:
                    last_name = '' 

            # ========== EMAIL ==========
            email_raw = str(row.get('Email') or '').strip().lower()
            if not email_raw:
                # Generate placeholder email for Simple format
                if template_format == 'SIMPLE':
                    email = f"{employee_code_normalized.lower()}@placeholder.local"
                    result['warnings'].append(
                        f"Row {idx}: No email → using placeholder '{email}'. Update later."
                    )
                else:
                    result['errors'].append(f"Row {idx}: Email is required")
                    continue
            else:
                if not _validate_email(email_raw):
                    result['errors'].append(f"Row {idx}: Invalid email '{email_raw}'")
                    continue
                email = email_raw

            # ========== PHONE ==========
            phone = str(row.get('Phone') or '').strip()
            if not phone:
                if template_format == 'SIMPLE':
                    phone = PLACEHOLDER_PHONE
                else:
                    result['errors'].append(f"Row {idx}: Phone is required")
                    continue

            # ========== DATES ==========
            if template_format == 'FULL':
                date_of_joining = _parse_date(row.get('Date of Joining'), 'Date of Joining', idx)
                date_of_birth = _parse_date(row.get('Date of Birth'), 'Date of Birth', idx)
            else:
                # Simple format uses placeholders
                date_of_joining = _parse_date(row.get('Date of Joining'), 'DOJ', idx, default=PLACEHOLDER_DOJ)
                date_of_birth = _parse_date(row.get('Date of Birth'), 'DOB', idx, default=PLACEHOLDER_DOB)

            # ========== OPTIONAL FIELDS ==========
            gender = _parse_gender(row.get('Gender'))
            personal_email = str(row.get('Personal Email') or '').strip().lower() or None

            # Status parsing (Simple format uses "Active"/"Inactive")
            is_active, status_value = _parse_status(row.get('Status'))
            if not is_active:
                status_value = 'TERMINATED'

            # Department
            department_name = str(row.get('Department') or '').strip()
            structure_location = None
            if department_name:
                structure_location = departments_cache.get(department_name.lower())
                if not structure_location:
                    result['warnings'].append(
                        f"Row {idx}: Department '{department_name}' not found. "
                        f"Employee created without department. Create dept in Settings → Departments."
                    )

            # Position
            position_title = str(row.get('Position') or '').strip()
            position = None
            if position_title:
                position = positions_cache.get(position_title.lower())
                if not position:
                    result['warnings'].append(
                        f"Row {idx}: Position '{position_title}' not found. Employee created without position."
                    )

            # ========== MATCH EXISTING EMPLOYEE ==========
            matched_employee = None
            for emp in Employee.objects.filter(is_deleted=False).only(
                'id', 'employee_id', 'official_email'
            ):
                if _normalize_employee_code(emp.employee_id) == employee_code_normalized:
                    matched_employee = emp
                    break
                if emp.official_email.lower() == email:
                    matched_employee = emp
                    break

            if matched_employee:
                if skip_existing:
                    result['skipped'] += 1
                    continue
                # Update
                matched_employee.first_name = first_name
                matched_employee.last_name = last_name
                matched_employee.official_email = email
                matched_employee.personal_email = personal_email
                matched_employee.phone_number = phone
                matched_employee.date_of_birth = date_of_birth
                matched_employee.date_of_joining = date_of_joining
                matched_employee.status = status_value
                if gender:
                    matched_employee.gender = gender
                if position:
                    matched_employee.position = position
                if structure_location:
                    matched_employee.structure_location = structure_location
                matched_employee.save()
                result['updated'] += 1
            else:
                # Create new
                new_emp = Employee.objects.create(
                    employee_id=employee_code_raw,  # Preserve original (NL001)
                    first_name=first_name,
                    last_name=last_name,
                    official_email=email,
                    personal_email=personal_email,
                    phone_number=phone,
                    date_of_birth=date_of_birth,
                    date_of_joining=date_of_joining,
                    gender=gender,
                    position=position,
                    structure_location=structure_location,
                    status=status_value,
                )
                result['created'] += 1
                result['created_employees'].append({
                    'employee_id': new_emp.employee_id,
                    'full_name': new_emp.full_name,
                    'email': new_emp.official_email,
                })

        except ValueError as ve:
            result['errors'].append(str(ve))
        except Exception as exc:
            result['errors'].append(f"Row {idx}: {str(exc)}")

    return result


def get_sample_csv_content() -> str:
    """Generate FULL sample CSV template."""
    headers = FULL_REQUIRED + FULL_OPTIONAL
    sample_rows = [
        ['NL001', 'Lince', '-', 'lince@company.com', '9876543210',
         '2024-01-15', '1995-06-20', 'Engineering', 'MALE', '', ''],
        ['NL002', 'Nisha', 'Yadav', 'nisha@company.com', '9876543211',
         '2024-02-10', '1996-03-15', 'Engineering', 'FEMALE', '', ''],
        ['NL003', 'Abhijith', 'Rajendran', 'abhijith@company.com', '9876543212',
         '2024-03-05', '1994-11-08', 'Engineering', 'MALE', '', ''],
    ]
    lines = [','.join(headers)]
    for row in sample_rows:
        escaped = [f'"{val}"' if ',' in str(val) else str(val) for val in row]
        lines.append(','.join(escaped))
    return '\n'.join(lines)


def get_simple_csv_content() -> str:
    """Generate SIMPLE sample CSV (matching attendance project format)."""
    headers = ['Employee Code', 'Employee Name', 'Department', 'Email', 'Status']
    sample_rows = [
        ['NL001', 'Lince', 'Admin', '', 'Active'],
        ['NL002', 'Nisha Yadav', 'Sales', '', 'Active'],
        ['NL003', 'Abhijith Rajendran', 'IT', '', 'Active'],
    ]
    lines = [','.join(headers)]
    for row in sample_rows:
        escaped = [f'"{val}"' if ',' in str(val) else str(val) for val in row]
        lines.append(','.join(escaped))
    return '\n'.join(lines)